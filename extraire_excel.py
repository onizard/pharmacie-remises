"""
Extraction et normalisation des données PDF Astera → Excel
Approche structurée : DCI + dosage + forme + quantité + PDA

Dépendances :
    pip install pdfplumber openpyxl rapidfuzz

Usage :
    python extraire_excel.py
"""

import re
from pathlib import Path
import json
from collections import defaultdict
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rapidfuzz import fuzz
from rapidfuzz.distance import Levenshtein as LevenshteinDist
from collections import Counter

PDF_DIR    = Path("pdfs_remises")
OUTPUT     = Path("remises_partenariat.xlsx")
BDPM_FILE  = Path("CIS_CIP_bdpm.txt")   # téléchargeable sur base-donnees-publique.medicaments.gouv.fr
PUHT_FILE  = Path("puht_astera.json")   # généré par scraper_puht.py

# ── Abréviations de labos à supprimer ────────────────────────────────────────
ABREV_LABOS = sorted([
    "ARROW", "ARR", "ARL",
    "BIOGARAN", "BIO", "BGR",
    "VIATRIS", "VIA", "MYL", "MYP",
    "PFIZER", "PFI",
    "SANDOZ", "SDZ", "SAN",
    "ZENTIVA", "ZEN",
    "TEVA", "TEV",
    "CRISTERS", "CRI",
    "ZYDUS", "ZYD",
    "CORREVIO", "CPH",
    "ABACUS",
    "WEGOVY", "STA",
    "EG", "GE", "ZTL", "REF", "SA", "QVL", "NOR", "KS", "QIL", "SUB",
], key=len, reverse=True)  # Plus longs d'abord pour éviter les conflits

# ── Unités de dosage ──────────────────────────────────────────────────────────
UNITES_DOSE = r'(?:MG/ML|MG|MCG|µG|UG|NG|G/ML|ML|UI/ML|MUI|MMOL|MOL|PC|%|G(?![A-Z]))'

# ── Formes pharmaceutiques ────────────────────────────────────────────────────
FORMES = r'(?:CPR|GELU|GELU|CAPS|COMP|AMP|SOL|PDR|CRE|GEL|POM|SUP|SPA|INJ|PERF|DISP|BUV|GTT|SACH|VERN|VERNIS|SPRAY|NAS|OPH|EAR|CPS|SEC|ORO|LP|LA|LI)'

# ── Extraction brute ──────────────────────────────────────────────────────────

def extraire_labo_fichier(nom_fichier: str) -> str:
    parts = nom_fichier.split(" - ")
    if len(parts) >= 2:
        labo = parts[1].strip()
        if any(x in labo for x in ["Liste", "CIP", "Partenariat"]):
            return re.sub(r'\bLABO\b', '', parts[0]).strip()
        return labo
    return nom_fichier

def nettoyer_taux(taux: str) -> float | None:
    if not taux:
        return None
    taux = taux.replace(",", ".").replace("%", "").replace(" ", "").strip()
    try:
        return float(taux)
    except ValueError:
        return None

def nettoyer_prix(prix: str) -> float | None:
    """Convertit '167,78€' en 167.78"""
    if not prix:
        return None
    prix = prix.replace(",", ".").replace("€", "").replace(" ", "").strip()
    try:
        return float(prix)
    except ValueError:
        return None

def extraire_pdf(pdf_path: Path) -> list[dict]:
    rows = []
    labo = extraire_labo_fichier(pdf_path.stem)
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 4:
                        continue
                    cip13    = (row[1] or "").strip()
                    libelle  = (row[2] or "").strip().upper()
                    taux_str = (row[3] or "").strip()
                    if cip13 == "CIP/ACL 13" or not cip13:
                        continue
                    if not re.fullmatch(r"\d{13}", cip13):
                        continue
                    # 5ème colonne : parfois PU HT (WEGOVY), parfois prix net direct (ABACUS)
                    puht_str = (row[4] if len(row) > 4 else "").strip() if row else ""
                    taux_val = nettoyer_taux(taux_str)
                    puht_val = nettoyer_prix(puht_str)
                    # Pas de taux mais un prix en €  → prix net fourni directement par le labo
                    punet_pdf = puht_val if (taux_val is None and puht_val is not None) else None
                    rows.append({
                        "Labo":         labo,
                        "CIP13":        cip13,
                        "Libellé brut": libelle,
                        "RSF %":        taux_val,
                        "PU NET pdf":   punet_pdf,
                    })
    return rows

# ── Parseur structuré ─────────────────────────────────────────────────────────

# Molécules abrégées à protéger
MOLECULES_PROTEGEES = {'H', 'A'}  # H=hydrochlorothiazide, A=acide clavulanique

DCI_CORRECTIONS = {
    r'\bCEFTRIAXIONE\b':    'CEFTRIAXONE',
    r'\bCLARITHROMYCYNE\b': 'CLARITHROMYCINE',
}

def supprimer_abrev(libelle: str) -> str:
    """Supprime les abréviations de labos, collées ou non au dosage."""
    for abrev in ABREV_LABOS:
        if abrev.upper() in MOLECULES_PROTEGEES:
            continue  # Protège H et A
        pattern = r'(?<=\s)' + re.escape(abrev) + r'(?=\s|\d|$)'
        libelle = re.sub(pattern, '', libelle, flags=re.IGNORECASE)
    # Supprime artefacts isolés entre espaces (PAS H = hydrochlorothiazide)
    for artefact in ['SEC', 'ORO', 'QUI', 'DISP', 'SP', 'TB']:
        libelle = re.sub(r'(?<![A-Z])\b' + artefact + r'\b(?![A-Z])', '', libelle, flags=re.IGNORECASE)
    # X seul (pas en début, pas collé à lettres)
    libelle = re.sub(r'(?<= )X(?= |$)', '', libelle)
    # Corrections orthographiques de DCI
    for pattern, correction in DCI_CORRECTIONS.items():
        libelle = re.sub(pattern, correction, libelle, flags=re.IGNORECASE)
    # Supprime INJ quand IV est également présent (redondant)
    if re.search(r'\bIV\b', libelle, re.IGNORECASE):
        libelle = re.sub(r'\bINJ\b\s*', '', libelle, flags=re.IGNORECASE)
    return re.sub(r' {2,}', ' ', libelle).strip()

def traiter_combinaisons(libelle: str) -> str:
    """
    Traite les associations de molécules avec dosages implicites.
    Appelé sur le libellé brut AVANT suppression des abréviations labo.
    """
    lib = libelle.upper().strip()

    # ── BISOPROLOL H → BISOPROLOL/HYD dose/6.25MG ───────────────────────────
    def norm_dose(dose):
        return re.sub(r'(\d+)MG(\d)\b(?!\d)', lambda x: f"{x.group(1)}.{x.group(2)}MG", dose)

    m = re.search(r'\bBISOPROLOL\s+H\b.*?(\d+(?:\.\d+)?MG\d?)\b', lib, re.IGNORECASE)
    if m:
        dose = norm_dose(m.group(1))
        reste = lib[m.end():]
        return f"BISOPROLOL/HYD {dose}/6.25MG{reste}"

    # ── BISOPROLOL dosage décimal en deux parties : 3MG 75CPR 30SEC → 3.75MG 30CPR ─
    m = re.search(r'\bBISOPROLOL\s+(\d+)MG\s+(\d{1,2})(CPR|GELU|COMP)\s+(\d+)\s*(?:SEC)?\b', lib, re.IGNORECASE)
    if m:
        dose = f'{m.group(1)}.{m.group(2)}MG'
        qty  = m.group(4)
        form = m.group(3).upper()
        reste = lib[m.end():]
        return f'BISOPROLOL {dose} {qty}{form}{reste}'

    # ── AMOXICILLINE A → AMOX/CLAV ───────────────────────────────────────────
    if re.search(r'\bAMOXICILLINE\s+A\b', lib, re.IGNORECASE):
        # Cas PDR EN/NN XML (suspension)
        m_pdr = re.search(r'\bAMOXICILLINE\s+A\b\s+(?:\S+\s+)*PDR\s+(?:EN|NN)\s*(\d+)\s*ML', lib, re.IGNORECASE)
        if m_pdr:
            return f"AMOX/CLAV PDR {m_pdr.group(1)}ML"
        # Cas avec dosage explicite
        RATIO = {'125MG': '31.25MG', '250MG': '62.5MG', '500MG': '62.5MG',
                 '875MG': '125MG', '1G': '125MG'}
        for amox, clav in RATIO.items():
            if amox in lib:
                idx = lib.find(amox) + len(amox)
                reste = lib[idx:]
                return f"AMOX/CLAV {amox}/{clav}{reste}"
        # Fallback
        lib = re.sub(r'\bAMOXICILLINE\s+A\b', 'AMOX/CLAV', lib, flags=re.IGNORECASE)
        # Inférence du dosage quand absent du PDF
        if re.search(r'\b(?:8|12)\s*(?:SACH|SAC|S)\b', lib, re.IGNORECASE):
            lib = lib.replace('AMOX/CLAV', 'AMOX/CLAV 1G/125MG', 1)
        elif re.search(r'\b(?:16|24)\s*CPR\b', lib, re.IGNORECASE):
            lib = lib.replace('AMOX/CLAV', 'AMOX/CLAV 500MG/62.5MG', 1)

    return lib

def separer_molecules_collees(libelle: str) -> str:
    """
    Insère un espace entre deux noms de molécules collés.
    Ex: BRINZOLAMIDETIM → BRINZOLAMIDE TIM
        CANDESARTANH → CANDESARTAN H
        AMLODIPVALS → AMLODIP VALS
    """
    # Suffixes de molécules couramment collés
    suffixes = [
        'TIM',   # timolol
        'VALS',  # valsartan
        'HCTZ',  # hydrochlorothiazide
    ]
    for suf in suffixes:
        libelle = re.sub(r'([A-Z])' + suf + r'\b', r'\1 ' + suf, libelle)

    # Cas H collé : molécule se terminant par une consonne + H + chiffre ou espace
    # Ex: CANDESARTANH 8MG → CANDESARTAN H 8MG
    # Mais pas: 2MG5H (dosage)
    libelle = re.sub(r'([A-Z]{4,})H(?=\s|\d)', r'\1 H', libelle)

    return libelle

def normaliser_bt(libelle: str) -> str:
    """
    Convertit BT[N] en N+FORME en détectant la forme présente dans le libellé.
    Corrige aussi l'ordre FORME N → NFORME.
    Ex: SACH BT24 → 24SACH, BT60 GELU → 60GELU, BT28CPR → 28CPR
    """
    # Normalise abréviations de sachet → SACH (ex: 12S, 8SAC, 12 SAC → 12SACH)
    libelle = re.sub(r'\b(\d+)\s*SAC\b', r'\1SACH', libelle, flags=re.IGNORECASE)
    libelle = re.sub(r'\b(\d+)S\b', r'\1SACH', libelle, flags=re.IGNORECASE)
    # GLE (granulé) → PDR
    libelle = re.sub(r'\bGLE\b', 'PDR', libelle, flags=re.IGNORECASE)
    # AM → AMP (ampoule tronquée dans certains PDFs : 1AM → 1AMP)
    libelle = re.sub(r'\b(\d+)AM\b', r'\1AMP', libelle, flags=re.IGNORECASE)

    FORMES_LIST = ['GELU', 'CAPS', 'COMP', 'SUPP', 'SACH', 'VERN', 'VERNIS',
                   'SPRAY', 'PERF', 'DISP', 'CPR', 'AMP', 'SOL', 'PDR',
                   'CRE', 'GEL', 'POM', 'SUP', 'SPA', 'INJ', 'BUV', 'GTT']
    FORMES_RE = '|'.join(FORMES_LIST)

    # Cherche la forme présente dans le libellé (hors BT)
    lib_sans_bt = re.sub(r'\bBT\s*\d+\b', '', libelle, flags=re.IGNORECASE)
    forme_match = re.search(r'\b(' + FORMES_RE + r')\b', lib_sans_bt, re.IGNORECASE)
    forme = forme_match.group(1).upper() if forme_match else 'CPR'

    # Cas 1 : BT + nombre + forme collée → retire juste BT
    libelle = re.sub(r'\bBT\s*(\d+)\s*(' + FORMES_RE + r')\b', r'\1\2', libelle, flags=re.IGNORECASE)
    # Cas 2 : BT + nombre seul → N + forme détectée
    libelle = re.sub(r'\bBT\s*(\d+)\b', lambda m: f"{m.group(1)}{forme}", libelle, flags=re.IGNORECASE)

    # Corrige ordre FORME N → NFORME, même si N est suivi de SEC (sécable)
    libelle = re.sub(
        r'\b(' + FORMES_RE + r')\s+(\d+)(?:SEC)?\b(?!\s*(?:MG|G(?![A-Z])|ML|UI|MCG|PC|%))',
        lambda m: f"{m.group(2)}{m.group(1).upper()}",
        libelle, flags=re.IGNORECASE
    )
    # Supprime la forme orpheline si elle est dupliquée (ex: "SACH 24SACH" → "24SACH")
    libelle = re.sub(r'\b(' + FORMES_RE + r')\s+(\d+\1)\b', r'\2', libelle, flags=re.IGNORECASE)

    return re.sub(r' {2,}', ' ', libelle).strip()

def normaliser_concentration(libelle: str) -> str:
    """Convertit 'N PC X ML' en 'N*10 MG/ML' pour uniformiser les concentrations.
    Ex: 2PC 5ML → 20MG/ML  (car 2% de 1ml = 20mg/ml)
    """
    def convertir(m):
        pct = float(m.group(1).replace(',', '.'))
        mg_ml = pct * 10
        val = int(mg_ml) if mg_ml == int(mg_ml) else mg_ml
        return f"{val}MG/ML"

    return re.sub(
        r'(\d+(?:[.,]\d+)?)\s*(?:PC|%)\s+\d+(?:[.,]\d+)?\s*ML',
        convertir,
        libelle,
        flags=re.IGNORECASE
    )

def inserer_espaces(libelle: str) -> str:
    """Insère les espaces manquants entre dose/unité et quantité."""
    # Sépare forme/qualificatif collé à un dosage (ex: COLLY0MG1 → COLLY 0MG1)
    libelle = re.sub(r'\b(COLLY|COL|INJ|GTT)(\d)', r'\1 \2', libelle, flags=re.IGNORECASE)
    # NMG75/NMG25 → N.75MG/N.25MG (2 chiffres décimaux : 3MG75, 1MG25, 0MG25…)
    libelle = re.sub(r'\b(\d+)MG(\d{2})\b', lambda m: f'{m.group(1)}.{m.group(2)}MG', libelle)
    # NMG5 → N.5MG (1 chiffre décimal : 2MG5, 7MG5…)
    libelle = re.sub(
        r'\b(\d+)MG(\d)\b(?!\d)',
        lambda m: f'{m.group(1)}.{m.group(2)}MG',
        libelle
    )
    # 0.NMG NML → 0.NMG/ML NML (concentration dans un flacon : /ML manquant dans le PDF)
    libelle = re.sub(
        r'\b(0\.\d+)MG\s+(\d+)ML\b',
        r'\1MG/ML \2ML',
        libelle, flags=re.IGNORECASE
    )
    # Insère espace entre unité de dose et chiffre de quantité collés
    # Ex: 500MG60CPR → 500MG 60CPR, 5600UI12CPR → 5600UI 12CPR
    # Lookahead sur la forme pour éviter de couper un vrai dosage
    libelle = re.sub(
        r'(\d+(?:[.,]\d+)?' + UNITES_DOSE + r')(\d+(?=' + FORMES + r'))',
        r'\1 \2',
        libelle, flags=re.IGNORECASE
    )
    return libelle

def parser_libelle(libelle: str) -> dict:
    """
    Extrait les composants structurés d'un libellé.
    Retourne: {dci, pda, dosage, quantite, forme, suffixe}
    """
    lib = libelle.strip().upper()

    # 1. Détection PDA
    pda = bool(re.search(r'\bPDA\b', lib))
    lib = re.sub(r'\bPDA\b', '', lib).strip()

    # 2. Suppression abréviations labo
    lib = supprimer_abrev(lib)

    # 2a. Combinaisons de molécules (avant suppression abréviations)
    lib = traiter_combinaisons(lib)

    # 2a2. Séparation des molécules collées
    lib = separer_molecules_collees(lib)

    # 2b. Normalisation concentrations NPC XML → N*10MG/ML
    lib = normaliser_concentration(lib)

    # 2c. Normalisation BT[N] → NCPR
    lib = normaliser_bt(lib)

    # 3. Insertion espaces manquants
    lib = inserer_espaces(lib)
    lib = re.sub(r' {2,}', ' ', lib).strip()

    # 4. Extraction dosage(s) : peut être composite ex: 10/160MG, 5600UI, 2PC, 0.25MG
    # Dosage composite : 10/160MG ou 10/160 (sans unité explicite après le slash)
    # Dosage : capture N[/N][unité] ou N/N sans unité (ex: 10/160)
    # UNITES_DOSE couvre MG, G, ML, UI, %, etc.
    # Dosage composite NUnit/NUnit en premier (ex: 1G/125MG, 875MG/125MG)
    _U = r'(?:MG/ML|MG|MCG|ML|UI|G(?![A-Z]))'
    pattern_dose = (r'(\d+(?:[.,]\d+)?' + _U + r'(?:/' + r'\d+(?:[.,]\d+)?' + _U + r')+'
                    r'|\d+(?:[.,]\d+)?(?:/\d+(?:[.,]\d+)?)*\s*(?:MG/ML|UI/ML|MUI|MG|MCG|ML|MMOL|MOL|PC|UI|%|G(?![A-Z]))'
                    r'|\d+/\d+(?:[.,]\d+)?)')
    doses = re.findall(pattern_dose, lib, re.IGNORECASE)
    dosage = ' '.join(d.strip() for d in doses) if doses else ''

    # 5. Extraction quantité : nombre suivi d'une forme
    pattern_qte = r'(\d+)\s*(' + FORMES + r'(?:\s+' + FORMES + r')*)'
    qte_match = re.search(pattern_qte, lib, re.IGNORECASE)
    quantite = ''
    forme    = ''
    if qte_match:
        quantite = qte_match.group(1)
        forme    = qte_match.group(2).strip().upper()

    # 6. DCI = tout ce qui précède le premier dosage ou la quantité
    dci = lib
    # Retire dosage
    for d in doses:
        dci = dci.replace(d, '')
    # Retire quantité+forme
    if qte_match:
        dci = dci[:dci.find(qte_match.group(0))] if qte_match.group(0) in dci else dci
    dci = re.sub(r'\d+', '', dci)           # retire chiffres résiduels
    dci = re.sub(r'[.,]', '', dci)          # retire . et ,
    dci = re.sub(r'(?<![A-Z])/|/(?![A-Z])', '', dci)  # retire / isolés, préserve AMOX/CLAV
    dci = re.sub(r'\s+', ' ', dci).strip()

    # 7. Suffixe (SEC, DISP, ORO, LP, etc.) = mots après la forme/quantité
    suffixe = ''
    if qte_match:
        reste = lib[qte_match.end():].strip()
        reste = re.sub(r'\s+', ' ', reste).strip()
        # Supprime les qualificatifs inutiles
        for mot in ['SEC', 'DISP', 'SP', 'ORO', 'QUI', 'X', 'TB']:
            reste = re.sub(r'\b' + mot + r'\b', '', reste, flags=re.IGNORECASE)
        suffixe = re.sub(r'\s+', ' ', reste).strip()

    return {
        'dci':      dci,
        'pda':      pda,
        'dosage':   dosage.upper(),
        'quantite': quantite,
        'forme':    forme.upper(),
        'suffixe':  suffixe.upper(),
    }

def cle_normalisation(parsed: dict) -> str:
    """Clé canonique pour regrouper les références identiques."""
    pda_tag = '_PDA' if parsed['pda'] else ''
    return f"{parsed['dci']}|{parsed['dosage']}|{parsed['quantite']}|{parsed['forme']}|{parsed['suffixe']}{pda_tag}"

def construire_libelle_normalise(parsed: dict) -> str:
    """Reconstruit un libellé propre et standardisé."""
    parts = [parsed['dci']]
    if parsed['pda']:
        parts.append('PDA')
    if parsed['dosage']:
        parts.append(parsed['dosage'])
    if parsed['quantite'] and parsed['forme']:
        parts.append(f"{parsed['quantite']}{parsed['forme']}")
    if parsed['suffixe']:
        parts.append(parsed['suffixe'])
    return ' '.join(p for p in parts if p)

# ── Correction automatique des fautes d'orthographe DCI ──────────────────────

def corriger_dci_typos(all_rows: list) -> dict:
    """
    Détecte les variantes orthographiques d'un même DCI (ex: CLARITHROMYCYNE /
    CLARITHROMYCINE) et retourne un dict {variante_rare: forme_canonique}.

    Deux DCI sont considérés variantes si :
      - ratio de similarité ≥ 92 % (≈ 1-2 caractères différents)
      - la différence de longueur est ≤ 2
      - ils partagent au moins un dosage ou une forme pharmaceutique
    Le nom le plus fréquent dans le dataset est retenu comme forme canonique.
    """
    dci_count = Counter(row['_parsed']['dci'] for row in all_rows)
    dcis = [d for d in dci_count if d]  # exclut chaînes vides

    dci_contexts: dict = defaultdict(set)
    for row in all_rows:
        p = row['_parsed']
        if p['dci']:
            dci_contexts[p['dci']].add((p['dosage'], p['forme']))

    # Union-Find : le représentant d'un groupe est toujours le plus fréquent
    parent = {d: d for d in dcis}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra == rb:
            return
        # Le moins fréquent pointe vers le plus fréquent
        if dci_count[ra] >= dci_count[rb]:
            parent[rb] = ra
        else:
            parent[ra] = rb

    for i, d1 in enumerate(dcis):
        for d2 in dcis[i + 1:]:
            w1, w2 = d1.split(), d2.split()
            # La faute doit être UNIQUEMENT dans le premier mot (la molécule elle-même).
            # Les mots suivants (qualificatifs : NEC, IV, IM, LP, A, H…) doivent être
            # identiques — sinon c'est un produit différent, pas une faute d'orthographe.
            # Ex. valide : CLARITHROMYCYNE vs CLARITHROMYCINE (1 mot, 1 char différent)
            # Ex. rejeté : CEFTRIAXONE NEC IV vs NEC IM (suffixes différents → produits ≠)
            #              TELMISARTAN A vs TELMISARTAN H (A=amlodipine ≠ H=hydrochlorothiaz.)
            #              TAMSULOSINE LP vs TAMSULOSINE CP (LP≠CP)
            if len(w1) != len(w2):
                continue
            if w1[1:] != w2[1:]:          # suffixes différents → produits distincts
                continue
            if LevenshteinDist.distance(w1[0], w2[0]) != 1:
                continue
            # Partagent-ils un dosage ou une forme ?
            ctx1, ctx2 = dci_contexts[d1], dci_contexts[d2]
            if ({c[0] for c in ctx1} & {c[0] for c in ctx2}
                    or {c[1] for c in ctx1} & {c[1] for c in ctx2}):
                union(d1, d2)

    corrections = {}
    groups: dict = defaultdict(list)
    for d in dcis:
        groups[find(d)].append(d)
    for root, members in groups.items():
        if len(members) == 1:
            continue
        canonical = max(members, key=lambda d: dci_count[d])
        for m in members:
            if m != canonical:
                corrections[m] = canonical
    return corrections


# ── Prix PU HT Astera (scraper_puht.py) ──────────────────────────────────────

def charger_puht() -> dict:
    """Charge puht_astera.json → {cip13: puht_float}. Généré par scraper_puht.py."""
    if not PUHT_FILE.exists():
        return {}
    return json.loads(PUHT_FILE.read_text(encoding="utf-8"))


# ── Synonymes de libellés validés manuellement ────────────────────────────────

SYNONYMS_FILE = Path("libelle_synonyms.json")

def charger_synonymes() -> dict:
    """Charge libelle_synonyms.json → {libelle_source: libelle_cible}."""
    if not SYNONYMS_FILE.exists():
        return {}
    return json.loads(SYNONYMS_FILE.read_text(encoding="utf-8"))

def appliquer_synonymes(all_rows: list, synonymes: dict) -> int:
    """
    Pour chaque ligne dont le libellé normalisé est une clé de synonymes,
    remplace _parsed et _cle par ceux du libellé cible.
    Retourne le nombre de lignes remappées.
    """
    if not synonymes:
        return 0
    # Pré-calcule (_parsed, _cle) pour chaque libellé cible
    target_cache: dict = {}
    for target_lib in set(synonymes.values()):
        parsed = parser_libelle(target_lib)
        target_cache[target_lib] = (parsed, cle_normalisation(parsed))

    nb = 0
    for row in all_rows:
        current_lib = construire_libelle_normalise(row['_parsed'])
        if current_lib in synonymes:
            target_lib = synonymes[current_lib]
            parsed, cle = target_cache[target_lib]
            row['_parsed'] = dict(parsed)   # copie pour éviter les partages
            row['_cle']    = cle
            nb += 1
    return nb


# ── Enrichissement BDPM (prix publics + taux de remboursement) ───────────────

def charger_bdpm() -> dict:
    """
    Charge CIS_CIP_bdpm.txt et retourne {cip13: {'ppttc': float|None, 'taux_remb': str}}.

    Fichier téléchargeable (onglet « Téléchargement ») sur :
        https://base-donnees-publique.medicaments.gouv.fr/
    → choisir « CIS_CIP_bdpm.txt », placer dans le dossier de travail.

    Colonnes du fichier (séparateur tab, encodage latin-1) :
        0:CIS  1:CIP7  2:Libellé  3:Statut  4:État comm.  5:Date décl.
        6:CIP13  7:Agrément coll.  8:Taux remb.  9:PPTTC  10:Indications
    """
    if not BDPM_FILE.exists():
        return {}
    bdpm = {}
    with open(BDPM_FILE, encoding='latin-1') as f:
        for line in f:
            cols = line.rstrip('\n').split('\t')
            if len(cols) < 10:
                continue
            cip13 = cols[6].strip()
            if not re.fullmatch(r'\d{13}', cip13):
                continue
            ppttc_str = cols[9].strip().replace(',', '.')
            try:
                ppttc = float(ppttc_str) if ppttc_str else None
            except ValueError:
                ppttc = None
            bdpm[cip13] = {
                'ppttc':     ppttc,
                'taux_remb': cols[8].strip(),
            }
    return bdpm


# ── Mise en forme Excel ───────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="DCE6F1")
BORDER_SIDE = Side(style="thin", color="B8CCE4")
CELL_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)

def style_header(cell):
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = CELL_BORDER

def style_data(cell, alt=False, center=False):
    cell.font      = DATA_FONT
    cell.border    = CELL_BORDER
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    if alt:
        cell.fill = ALT_FILL

def ecrire_lignes(ws, rows, headers, start_row=2):
    for r_idx, row in enumerate(rows, start_row):
        alt = r_idx % 2 == 0
        for col, key in enumerate(headers, 1):
            val  = row.get(key)
            cell = ws.cell(row=r_idx, column=col, value=val)
            center = key in ("CIP13", "RSF %", "Labo", "PU HT", "PU NET")
            style_data(cell, alt, center)
            if key == "RSF %":
                cell.number_format = '0.00"%"'
            elif key == "CIP13":
                cell.number_format = "@"
            elif key in ("PU HT", "PU NET") and val is not None:
                cell.number_format = '#,##0.00 "€"'
                cell.alignment = Alignment(horizontal="center", vertical="center")

def appliquer_largeurs(ws):
    ws.column_dimensions["A"].width = 12   # Labo
    ws.column_dimensions["B"].width = 18   # CIP13
    ws.column_dimensions["C"].width = 52   # Libellé
    ws.column_dimensions["D"].width = 12   # PU HT
    ws.column_dimensions["E"].width = 12   # RSF %
    ws.column_dimensions["F"].width = 12   # PU NET
    ws.freeze_panes = "A2"

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    pdfs = sorted(PDF_DIR.glob("*CIP*.pdf"))
    if not pdfs:
        print("❌  Aucun PDF trouvé dans", PDF_DIR)
        return

    print(f"📄  {len(pdfs)} PDF(s) à traiter...\n")

    # 0. Chargement BDPM (optionnel : prix publics + taux de remboursement)
    puht_map = charger_puht()
    if puht_map:
        print(f"💶  {len(puht_map)} prix PU HT chargés ({PUHT_FILE.name})\n")
    else:
        print(f"ℹ️   {PUHT_FILE.name} non trouvé — colonne PU HT vide.\n"
              f"    → Lancer scraper_puht.py pour récupérer les prix depuis Astera.\n")

    # 1. Extraction brute
    all_rows = []
    for pdf_path in pdfs:
        labo = extraire_labo_fichier(pdf_path.stem)
        print(f"  ⚙️   {pdf_path.name}  →  Labo : {labo}")
        rows = extraire_pdf(pdf_path)
        print(f"       {len(rows)} références extraites")
        all_rows.extend(rows)

    print(f"\n🔧  Parsing et normalisation structurée...", flush=True)

    # 2. Parse chaque libellé
    for row in all_rows:
        row['_parsed'] = parser_libelle(row['Libellé brut'])
        row['_cle']    = cle_normalisation(row['_parsed'])

    # 2b. Correction automatique des fautes d'orthographe DCI
    corrections_dci = corriger_dci_typos(all_rows)
    if corrections_dci:
        print(f"  ✏️   {len(corrections_dci)} correction(s) DCI :")
        for wrong, right in sorted(corrections_dci.items()):
            print(f"      {wrong} → {right}")
        for row in all_rows:
            dci = row['_parsed']['dci']
            if dci in corrections_dci:
                row['_parsed']['dci'] = corrections_dci[dci]
        for row in all_rows:
            row['_cle'] = cle_normalisation(row['_parsed'])

    # 2b.5 Application des synonymes de libellés validés manuellement
    synonymes = charger_synonymes()
    if synonymes:
        nb_syn = appliquer_synonymes(all_rows, synonymes)
        if nb_syn:
            print(f"  🔗  {nb_syn} référence(s) remappée(s) via {SYNONYMS_FILE.name}")

    # 2c. Enrichissement PU HT depuis Astera (scraper_puht.py)
    for row in all_rows:
        row['PU HT'] = puht_map.get(row['CIP13'])

    # 2d. Propagation PU HT par libellé normalisé (pour les CIP sans prix Astera)
    cle_to_puht: dict[str, float] = {}
    for row in all_rows:
        if row['PU HT'] is not None:
            cle_to_puht.setdefault(row['_cle'], row['PU HT'])

    nb_propages = 0
    for row in all_rows:
        if row['PU HT'] is None and row['_cle'] in cle_to_puht:
            row['PU HT'] = cle_to_puht[row['_cle']]
            nb_propages += 1
    if nb_propages:
        print(f"  📋  {nb_propages} PU HT propagé(s) depuis des références au même libellé")

    # 2e. Calcul PU NET
    for row in all_rows:
        puht = row['PU HT']
        rsf  = row.get('RSF %')
        if puht is not None and rsf is not None:
            row['PU NET'] = round(puht * (1 + rsf / 100), 4)
        elif row.get('PU NET pdf') is not None:
            # Prix net fourni directement dans le PDF (ex. ABACUS)
            row['PU NET'] = row['PU NET pdf']
        else:
            row['PU NET'] = None

    # 3. Choix du libellé modèle par groupe (clé canonique)
    #    On choisit le libellé reconstruit le plus fréquent ou le plus court
    groupes = defaultdict(list)
    for row in all_rows:
        groupes[row['_cle']].append(row['_parsed'])

    modeles = {}
    for cle, parsed_list in groupes.items():
        candidats = [construire_libelle_normalise(p) for p in parsed_list]
        modeles[cle] = min(candidats, key=lambda x: (len(x), x))

    # 4. Applique le libellé modèle
    for row in all_rows:
        row['Libellé'] = modeles[row['_cle']]

    nb_bruts   = len(set(r['Libellé brut'] for r in all_rows))
    nb_normalises = len(set(r['Libellé'] for r in all_rows))
    print(f"  {nb_bruts} libellés bruts → {nb_normalises} libellés normalisés\n")

    # 5. Construction Excel — onglet unique "Tous les labos"
    print("📊  Création du fichier Excel...")
    wb = Workbook()
    # Ordre : Labo | CIP13 | Libellé | PU HT | RSF % | PU NET
    headers = ["Labo", "CIP13", "Libellé", "PU HT", "RSF %", "PU NET"]

    ws = wb.active
    ws.title = "Tous les labos"
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))
    ws.row_dimensions[1].height = 20
    ecrire_lignes(ws, all_rows, headers)
    appliquer_largeurs(ws)
    ws.auto_filter.ref = f"A1:F{len(all_rows)+1}"

    wb.save(OUTPUT)
    print(f"\n✅  {len(all_rows)} références au total")
    print(f"📁  Fichier créé : {OUTPUT.resolve()}")

if __name__ == "__main__":
    main()