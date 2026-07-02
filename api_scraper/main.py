"""
API FastAPI — Break-Pharma Scraper Service

Endpoints :
  GET  /health               → sanity check
  POST /connect/{connector}  → teste les identifiants, met à jour connected dans Supabase
  POST /run/{connector}      → lance le scraping en arrière-plan
  GET  /status/{job_id}      → retourne le statut du job

Authentification : Bearer token Supabase (JWT de l'utilisateur break-pharma.fr)
Connecteurs supportés : ospharm, digipharmacie
"""

import asyncio
import json
import os
import re as _re
import time
import urllib.request
import uuid
from concurrent.futures import ThreadPoolExecutor
from typing import Optional

from fastapi import BackgroundTasks, FastAPI, File, Form, Header, HTTPException, Path, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from supabase_client import (
    _get_state_sync,
    _patch_state_sync,
    get_user_creds_for,
    patch_conn_test,
    patch_connector_connected,
    patch_job_status,
    save_user_creds,
    verify_token,
)

GH_TOKEN = os.environ.get("GH_TOKEN", "")
GH_REPO  = "onizard/pharmacie-remises"
GH_WORKFLOW      = "scraper_ospharm.yml"
GH_DIGI_WORKFLOW = "scraper.yml"
GH_TEST_WORKFLOW = "test_connector.yml"
GH_FSE_WORKFLOW  = "scraper_fse.yml"

GMAIL_USER     = os.environ.get("GMAIL_USER", "pharmaciemontmagny@gmail.com")
GMAIL_APP_PASS = os.environ.get("GMAIL_APP_PASSWORD", "")


class ConnectBody(BaseModel):
    user: str
    password: str

# ── App ────────────────────────────────────────────────────────────────────────

app = FastAPI(title="Break-Pharma Scraper API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://break-pharma.fr",
        "https://www.break-pharma.fr",
        "https://onizard.github.io",
        "http://localhost:5500",
        "http://127.0.0.1:5500",
    ],
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["*"],
    allow_credentials=True,
)

SUPPORTED_CONNECTORS = {"ospharm", "digipharmacie", "concentrateur", "gmail"}

# ── Client-side log store (in-memory, last 200 entries) ───────────────────────

_client_logs: list[dict] = []
_CLIENT_LOG_MAX = 200

# ── Job store (in-memory) ──────────────────────────────────────────────────────

_jobs: dict[str, dict] = {}
JOB_TTL   = 3600
_executor = ThreadPoolExecutor(max_workers=3)


def _cleanup_jobs():
    cutoff = time.time() - JOB_TTL
    stale  = [jid for jid, j in _jobs.items() if j.get("created", 0) < cutoff]
    for jid in stale:
        del _jobs[jid]


def _extract_token(authorization: str) -> str:
    token = authorization.removeprefix("Bearer ").strip()
    if not token:
        raise HTTPException(status_code=401, detail="Token manquant")
    return token


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/client-log")
async def post_client_log(request: Request, authorization: str = Header(default="")):
    """Reçoit les logs console du navigateur (erreurs, warnings, infos)."""
    import datetime
    try:
        body = await request.json()
    except Exception:
        body = {}
    entry = {
        "ts":    datetime.datetime.utcnow().isoformat() + "Z",
        "level": body.get("level", "log"),
        "msg":   str(body.get("msg", ""))[:2000],
        "url":   str(body.get("url", ""))[:200],
        "line":  body.get("line"),
        "user":  body.get("user", ""),
    }
    _client_logs.append(entry)
    if len(_client_logs) > _CLIENT_LOG_MAX:
        del _client_logs[:-_CLIENT_LOG_MAX]
    print(f"[client-log] {entry['level'].upper()} {entry['msg'][:120]}")
    return {"ok": True}


@app.get("/client-log")
async def get_client_logs(n: int = 50, level: str = ""):
    """Retourne les n derniers logs client (filtrables par level=error|warn|log)."""
    logs = _client_logs[-n:]
    if level:
        logs = [l for l in logs if l["level"] == level]
    return {"count": len(logs), "logs": logs}



@app.post("/connect/{connector}")
async def connect_connector(
    background_tasks: BackgroundTasks,
    body: ConnectBody,
    connector: str = Path(...),
    authorization: str = Header(default=""),
):
    """Enregistre les identifiants, lance le test en arrière-plan, retourne immédiatement."""
    if connector not in SUPPORTED_CONNECTORS:
        raise HTTPException(status_code=400, detail=f"Connecteur inconnu : {connector}")

    token = _extract_token(authorization)
    try:
        user_id = await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    creds = {"user": body.user, "pass": body.password}
    background_tasks.add_task(_run_conn_test_async, user_id, connector, creds, token)
    return {"status": "testing"}


def _reset_digi_state(user_id: str, user_token: str = ""):
    """Vide le cache de factures + les stats Digi (re-scrape complet / synchro)."""
    from supabase_client import _get_state_sync, _patch_state_sync
    state = _get_state_sync(user_id, user_token=user_token) or {}
    state["digi_month_stats"] = {}
    state["escompte_stats"]   = {}
    state["mdl_stats"]        = {}
    vj = dict(state.get("verif_job") or {})
    vj["invoice_cache"] = {}
    state["verif_job"] = vj
    _patch_state_sync(user_id, state, user_token=user_token)


def _reset_fse_state(user_id: str, user_token: str = ""):
    """Vide les stats FSE (re-scrape complet : ré-applique l'appariement à jour)."""
    from supabase_client import _get_state_sync, _patch_state_sync
    state = _get_state_sync(user_id, user_token=user_token) or {}
    state["fse_month_stats"] = {}
    _patch_state_sync(user_id, state, user_token=user_token)


@app.post("/run/{connector}")
async def run_connector(
    background_tasks: BackgroundTasks,
    connector: str = Path(...),
    resync: bool = False,
    authorization: str = Header(default=""),
):
    """Lance le scraping. OSPHARM → GitHub Actions. DIGIPHARMACIE → local.
    resync=true (Digi) : vide le cache + les stats avant de relancer → re-scrape
    complet de l'historique (ré-applique les parsers à jour, ex : RDP par palier)."""
    token = _extract_token(authorization)
    try:
        user_id = await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    # grossiste_gmail : traité ici directement
    if connector == "grossiste_gmail":
        loop = asyncio.get_event_loop()
        try:
            result = await loop.run_in_executor(_executor, lambda: _run_grossiste_gmail_sync(user_id, user_token=token))
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
        return result

    # fse-export : déclenche le scraper FSE Banque via GitHub Actions. Doit être traité ICI car
    # la route dédiée /run/fse-export est masquée par cette route générique /run/{connector}
    # (FastAPI matche dans l'ordre de déclaration) → sinon "Connecteur inconnu : fse-export".
    if connector == "fse-export":
        if resync:
            # Synchro complète : on vide les stats FSE avant de relancer pour que le
            # ré-export ré-applique l'appariement des virements à jour.
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(_executor, lambda: _reset_fse_state(user_id, token))
        background_tasks.add_task(_dispatch_gh_fse, user_id, token)
        return {"status": "dispatched", "resync": resync}

    if connector not in SUPPORTED_CONNECTORS:
        raise HTTPException(status_code=400, detail=f"Connecteur inconnu : {connector}")

    try:
        creds = await get_user_creds_for(user_id, connector, user_token=token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lecture credentials: {e}")

    _cleanup_jobs()
    job_id = str(uuid.uuid4())

    if connector == "ospharm":
        background_tasks.add_task(_dispatch_gh_ospharm, user_id, token)
        return {"job_id": job_id, "mode": "github_actions"}

    # DIGIPHARMACIE : dispatch GitHub Actions (self-hosted, proxy résidentiel, camoufox)
    if resync:
        # Synchro complète : on vide cache + stats AVANT de dispatcher, pour que le
        # scraper re-traite tout l'historique avec les parsers à jour.
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(_executor, lambda: _reset_digi_state(user_id, token))
    background_tasks.add_task(_dispatch_gh_digi, user_id, token)
    return {"job_id": job_id, "mode": "github_actions", "resync": resync}


async def _dispatch_gh_digi(user_id: str, user_token: str = ""):
    """Déclenche scraper.yml sur GitHub Actions (self-hosted, proxy résidentiel)."""
    loop  = asyncio.get_event_loop()
    cur   = await loop.run_in_executor(None, lambda: _get_state_sync(user_id, user_token))
    if (cur.get("verif_job") or {}).get("status") == "running":
        print(f"  [gh-dispatch] verif_job déjà en cours pour {user_id[:8]} — dispatch ignoré")
        return
    await patch_job_status(user_id, "verif_job", "running",
                           "Job en attente de démarrage…", [], user_token=user_token)
    if not GH_TOKEN:
        await patch_job_status(user_id, "verif_job", "error",
                               "GH_TOKEN manquant sur le serveur — contacter l'admin", [],
                               user_token=user_token)
        return
    url  = f"https://api.github.com/repos/{GH_REPO}/actions/workflows/{GH_DIGI_WORKFLOW}/dispatches"
    body = json.dumps({"ref": "master", "inputs": {"user_id": user_id}}).encode()
    req  = urllib.request.Request(url, data=body, method="POST", headers={
        "Authorization": f"Bearer {GH_TOKEN}",
        "Accept":        "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "Content-Type":  "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            print(f"  [gh-dispatch] HTTP {r.status} — scraper digi déclenché pour {user_id[:8]}")
    except Exception as e:
        print(f"  [gh-dispatch] ERREUR digi: {e}")
        await patch_job_status(user_id, "verif_job", "error",
                               f"Impossible de lancer le workflow GitHub: {e}", [],
                               user_token=user_token)


async def _dispatch_gh_ospharm(user_id: str, user_token: str = ""):
    """Déclenche le workflow GitHub Actions scraper_ospharm.yml."""
    await patch_job_status(user_id, "ospharm_job", "running",
                           "Chargement des données en cours…", [], user_token=user_token)

    if not GH_TOKEN:
        await patch_job_status(user_id, "ospharm_job", "error",
                               "GH_TOKEN manquant sur le serveur — contacter l'admin", [],
                               user_token=user_token)
        return

    url  = f"https://api.github.com/repos/{GH_REPO}/actions/workflows/{GH_WORKFLOW}/dispatches"
    body = json.dumps({"ref": "master", "inputs": {"user_id": user_id}}).encode()
    req  = urllib.request.Request(url, data=body, method="POST", headers={
        "Authorization": f"Bearer {GH_TOKEN}",
        "Accept":        "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "Content-Type":  "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            print(f"  [gh-dispatch] HTTP {r.status} — workflow ospharm déclenché pour {user_id[:8]}")
    except Exception as e:
        print(f"  [gh-dispatch] ERREUR: {e}")
        await patch_job_status(user_id, "ospharm_job", "error",
                               f"Impossible de lancer le workflow GitHub: {e}", [],
                               user_token=user_token)


@app.post("/run/fse-export")
async def run_fse_export(
    background_tasks: BackgroundTasks,
    authorization: str = Header(default=""),
):
    """Déclenche le scraper FSE Banque via GitHub Actions."""
    token = _extract_token(authorization)
    try:
        user_id = await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))
    background_tasks.add_task(_dispatch_gh_fse, user_id, token)
    return {"status": "dispatched"}


async def _dispatch_gh_fse(user_id: str, user_token: str = ""):
    """Déclenche scraper_fse.yml sur GitHub Actions (dates auto : jan N-1 → aujourd'hui)."""
    import datetime as _dt
    today     = _dt.date.today()
    date_from = f"{today.year - 1}-01-01"
    date_to   = today.strftime("%Y-%m-%d")

    loop = asyncio.get_event_loop()
    try:
        state = await loop.run_in_executor(None, lambda: _get_state_sync(user_id, user_token))
        state["fse_job"] = {"status": "running", "message": f"Export FSE lancé ({date_from} → {date_to})…", "error": ""}
        await loop.run_in_executor(None, lambda: _patch_state_sync(user_id, state, user_token))
    except Exception:
        pass

    if not GH_TOKEN:
        try:
            state = await loop.run_in_executor(None, lambda: _get_state_sync(user_id, user_token))
            state["fse_job"] = {"status": "error", "message": "", "error": "GH_TOKEN manquant sur le serveur"}
            await loop.run_in_executor(None, lambda: _patch_state_sync(user_id, state, user_token))
        except Exception:
            pass
        return

    url  = f"https://api.github.com/repos/{GH_REPO}/actions/workflows/{GH_FSE_WORKFLOW}/dispatches"
    body = json.dumps({"ref": "master", "inputs": {
        "user_id":   user_id,
        "date_from": date_from,
        "date_to":   date_to,
    }}).encode()
    req = urllib.request.Request(url, data=body, method="POST", headers={
        "Authorization":        f"Bearer {GH_TOKEN}",
        "Accept":               "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "Content-Type":         "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            print(f"  [gh-dispatch] HTTP {r.status} — fse-export déclenché pour {user_id[:8]} ({date_from}→{date_to})")
    except Exception as e:
        print(f"  [gh-dispatch] ERREUR fse: {e}")
        try:
            state = await loop.run_in_executor(None, lambda: _get_state_sync(user_id, user_token))
            state["fse_job"] = {"status": "error", "message": "", "error": str(e)}
            await loop.run_in_executor(None, lambda: _patch_state_sync(user_id, state, user_token))
        except Exception:
            pass


@app.get("/status/{job_id}")
def get_status(job_id: str):
    job = _jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job inconnu ou expiré")
    return job


# ── Grossiste helpers ──────────────────────────────────────────────────────────

_GROSSISTE_LABO_MAP = {
    "biogaran": "BIOGARAN", "teva": "TEVA", "mylan": "MYLAN",
    "viatris": "VIATRIS", "zydus": "ZYDUS", "sandoz": "SANDOZ",
    "zentiva": "ZENTIVA", "arrow": "ARROW", "cristers": "CRISTERS",
    "eg labo": "EG LABO", "eg labs": "EG LABO", "evolupharm": "EVOLUPHARM",
    "ranbaxy": "RANBAXY", "actavis": "ACTAVIS", "aurobindo": "AUROBINDO",
    "intas": "INTAS", "almus": "ALMUS",
}

def _norm_grossiste_labo(raw: str) -> str:
    import re
    n = (raw or "").lower()
    for kw, canon in _GROSSISTE_LABO_MAP.items():
        if kw in n:
            return canon
    m = re.match(r"([A-Z][A-Z\-\']+)", (raw or "").strip())
    return m.group(1) if m else (raw or "").upper().split()[0] if raw else "?"


def _parse_grossiste_bytes(xlsx_bytes: bytes) -> dict:
    """Parse feuille 'Récap par mois' → {year-MM: [{labo, qty, total_ht, ca_brut,
    paliers: [{taux, qty, brut, remise, net}]}]}.

    Le justificatif répartiteur ventile chaque labo par 'Tx Rem' (= palier RSF :
    0 / 2,5 / 5 / 10 / 20 / 25 / 30 / 40). On conserve ce détail par palier
    (montant remise = RSF effectivement obtenu) en plus des totaux par labo.
    """
    import io, re, openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "Récap par mois" not in wb.sheetnames:
        return {}
    ws = wb["Récap par mois"]

    month_acc: dict[str, dict] = {}
    current_month = None
    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue
        cell0 = str(row[0] or "")
        if "Mois comptable" in cell0:
            m = re.search(r"(\d{4})\s+(\d{2})", cell0)
            if m:
                current_month = f"{m.group(1)}-{m.group(2)}"
                month_acc.setdefault(current_month, {})
            continue
        if current_month is None:
            continue
        # Cols : Rep/Dep | nom | Tx Rem | qtes | Mt Vente Brut HT | Montant remise | CA net HT
        rep_dep, labo_raw, taux, qty, ca_brut_raw, remise_raw, ca_net = (list(row) + [None]*7)[:7]
        if rep_dep == "Rep G" and labo_raw and qty:
            labo = _norm_grossiste_labo(labo_raw)
            acc  = month_acc[current_month].setdefault(
                labo, {"qty": 0, "total_ht": 0.0, "ca_brut": 0.0, "paliers": {}})
            q = int(qty or 0)
            b = float(ca_brut_raw or 0)
            r = float(remise_raw or 0)
            n = float(ca_net or 0)
            acc["qty"]      += q
            acc["total_ht"] += n
            acc["ca_brut"]  += b
            try:
                tx = round(float(taux), 2)
            except (TypeError, ValueError):
                tx = None
            if tx is not None:
                p = acc["paliers"].setdefault(tx, {"qty": 0, "brut": 0.0, "remise": 0.0, "net": 0.0})
                p["qty"]    += q
                p["brut"]   += b
                p["remise"] += r
                p["net"]    += n

    wb.close()
    return {
        mk: sorted(
            [{"labo": l, "qty": d["qty"], "total_ht": round(d["total_ht"], 2), "ca_brut": round(d["ca_brut"], 2),
              "paliers": sorted(
                  [{"taux": tx, "qty": p["qty"], "brut": round(p["brut"], 2),
                    "remise": round(p["remise"], 2), "net": round(p["net"], 2)}
                   for tx, p in d["paliers"].items()],
                  key=lambda x: x["taux"])}
             for l, d in labos.items()],
            key=lambda r: r["labo"],
        )
        for mk, labos in sorted(month_acc.items())
    }


def _merge_paliers(a: list, b: list) -> list:
    """Fusion additive de deux listes de paliers [{taux, qty, brut, remise, net}]."""
    acc: dict = {}
    for src in (a or []), (b or []):
        for p in src:
            tx = p.get("taux")
            d  = acc.setdefault(tx, {"qty": 0, "brut": 0.0, "remise": 0.0, "net": 0.0})
            d["qty"]    += p.get("qty", 0)
            d["brut"]   += p.get("brut", 0)
            d["remise"] += p.get("remise", 0)
            d["net"]    += p.get("net", 0)
    return sorted(
        [{"taux": tx, "qty": d["qty"], "brut": round(d["brut"], 2),
          "remise": round(d["remise"], 2), "net": round(d["net"], 2)}
         for tx, d in acc.items()],
        key=lambda x: (x["taux"] if x["taux"] is not None else -1))


def _merge_grossiste_stats(existing: dict, new_stats: dict) -> dict:
    """Fusion additive : mois distincts → union ; mois communs → addition par labo
    (y compris la ventilation par palier)."""
    merged = dict(existing)
    for mk, new_rows in new_stats.items():
        if mk not in merged:
            merged[mk] = new_rows
        else:
            labo_map = {r["labo"]: dict(r) for r in merged[mk]}
            for nr in new_rows:
                if nr["labo"] in labo_map:
                    ex = labo_map[nr["labo"]]
                    ex["qty"]      += nr["qty"]
                    ex["total_ht"]  = round(ex["total_ht"] + nr["total_ht"], 2)
                    ex["ca_brut"]   = round(ex.get("ca_brut", 0) + nr.get("ca_brut", 0), 2)
                    ex["paliers"]   = _merge_paliers(ex.get("paliers"), nr.get("paliers"))
                else:
                    labo_map[nr["labo"]] = dict(nr)
            merged[mk] = sorted(labo_map.values(), key=lambda r: r["labo"])
    return merged


# ── Parse grossiste XLSX (depuis MinIO) ────────────────────────────────────────

class ParseGrossisteBody(BaseModel):
    storage_path: str  # chemin dans le bucket 'grossiste', ex: "user_id/ts_filename.xlsx"

@app.post("/parse/grossiste")
async def parse_grossiste(
    body: ParseGrossisteBody,
    authorization: str = Header(default=""),
):
    """Télécharge le XLSX grossiste depuis MinIO Storage, parse et sauvegarde grossiste_month_stats."""
    token = _extract_token(authorization)
    try:
        user_id = await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        _executor,
        lambda: _parse_grossiste_sync(user_id, body.storage_path)
    )
    return result


def _parse_grossiste_sync(user_id: str, storage_path: str) -> dict:
    from supabase_client import SUPA_URL, _supa_key
    _tok = _supa_key()
    HEADERS = {"apikey": _tok, "Authorization": f"Bearer {_tok}"}

    # 1. Télécharger le XLSX depuis Storage
    dl_url = f"{SUPA_URL}/storage/v1/object/grossiste/{storage_path}"
    req = urllib.request.Request(dl_url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=30) as r:
        xlsx_bytes = r.read()

    # 2. Parser
    grossiste_stats = _parse_grossiste_bytes(xlsx_bytes)
    if not grossiste_stats:
        raise HTTPException(status_code=422, detail="Aucune donnée extractible — vérifie le format du fichier.")

    # 3. Mettre à jour l'état
    from supabase_client import _get_state_sync, _patch_state_sync
    state = _get_state_sync(user_id) or {}
    state["grossiste_month_stats"] = _merge_grossiste_stats(
        state.get("grossiste_month_stats") or {}, grossiste_stats
    )
    _patch_state_sync(user_id, state)

    months   = sorted(grossiste_stats)
    total_q  = sum(r["qty"]      for rows in grossiste_stats.values() for r in rows)
    total_ht = sum(r["total_ht"] for rows in grossiste_stats.values() for r in rows)
    return {
        "status":  "done",
        "months":  months,
        "labos":   len({r["labo"] for rows in grossiste_stats.values() for r in rows}),
        "qty":     total_q,
        "total_ht": round(total_ht, 2),
        "grossiste_month_stats": grossiste_stats,
    }


# ── Upload direct XLSX grossiste (multipart, sans MinIO) ───────────────────────
# Le stockage self-hosted est du MinIO brut (auth S3) : supabase-js .upload() ne
# peut pas s'y authentifier côté navigateur. On reçoit donc le fichier en multipart
# direct (comme /parse/digi-pdf) et on parse en mémoire, sans passer par le bucket.
@app.post("/parse/grossiste-xlsx")
async def parse_grossiste_xlsx(
    file: UploadFile = File(...),
    reset: bool = Form(False),
    authorization: str = Header(default=""),
):
    token = _extract_token(authorization)
    try:
        user_id = await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    name = (file.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls")):
        raise HTTPException(status_code=422, detail="Fichier XLSX requis (.xlsx/.xls)")
    xlsx_bytes = await file.read()
    if len(xlsx_bytes) < 200:
        raise HTTPException(status_code=422, detail="Fichier trop court ou vide")

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        _executor, lambda: _parse_grossiste_xlsx_sync(user_id, xlsx_bytes, token, reset=reset)
    )


def _parse_grossiste_xlsx_sync(user_id: str, xlsx_bytes: bytes, user_token: str = "", reset: bool = False) -> dict:
    # Écriture de l'état via les helpers avec le TOKEN UTILISATEUR (RLS : l'user peut MAJ
    # sa propre ligne). La clé anon est en lecture seule sur user_state → 401 en écriture.
    # reset=True : on repart d'une base vide (ré-analyse : le 1er fichier réinitialise).
    from supabase_client import _get_state_sync, _patch_state_sync

    grossiste_stats = _parse_grossiste_bytes(xlsx_bytes)
    if not grossiste_stats:
        raise HTTPException(
            status_code=422,
            detail="Aucune donnée extractible — feuille 'Récap par mois' attendue dans le justificatif.",
        )

    state = _get_state_sync(user_id, user_token=user_token) or {}
    base  = {} if reset else (state.get("grossiste_month_stats") or {})
    state["grossiste_month_stats"] = _merge_grossiste_stats(base, grossiste_stats)
    _patch_state_sync(user_id, state, user_token=user_token)

    months   = sorted(grossiste_stats)
    total_q  = sum(r["qty"]      for rows in grossiste_stats.values() for r in rows)
    total_ht = sum(r["total_ht"] for rows in grossiste_stats.values() for r in rows)
    return {
        "status":   "done",
        "months":   months,
        "labos":    len({r["labo"] for rows in grossiste_stats.values() for r in rows}),
        "qty":      total_q,
        "total_ht": round(total_ht, 2),
        "grossiste_month_stats": state["grossiste_month_stats"],
    }


# ── Job grossiste Gmail — géré dans /run/{connector} ci-dessus ────────────────

def _run_grossiste_gmail_sync(user_id: str, user_token: str = "") -> dict:
    import imaplib
    import email as _email_lib
    from email.header import decode_header as _decode_header

    from supabase_client import _get_state_sync, _patch_state_sync, _get_connectors_sync

    # Credentials : user state en priorité, fallback vars d'env Render
    gmail_user = GMAIL_USER
    gmail_pass = GMAIL_APP_PASS
    try:
        conns = _get_connectors_sync(user_id, user_token)
        user_gmail = conns.get("gmail", {})
        if user_gmail.get("user") and user_gmail.get("pass"):
            gmail_user = user_gmail["user"]
            gmail_pass = user_gmail["pass"]
    except Exception:
        pass

    if not gmail_pass:
        raise ValueError("Credentials Gmail non configurés — renseignez-les dans CONNECTEURS → GMAIL.")

    # 1. Connexion IMAP Gmail (timeout 20s pour éviter un hang silencieux)
    mail = imaplib.IMAP4_SSL("imap.gmail.com", timeout=20)
    mail.login(gmail_user, gmail_pass)
    mail.select("INBOX")

    # 2. UIDs déjà traités
    state = _get_state_sync(user_id, user_token=user_token) or {}
    processed_uids: set = set(state.get("grossiste_gmail_uids") or [])

    # 3. Chercher les emails CERP Rouen depuis 2024 (sujet contient é → search par expéditeur)
    _, search_data = mail.search(None, 'FROM "cerp-rouen.fr" SINCE "01-Jan-2024"')
    candidate_uids = [int(u) for u in (search_data[0].split() if search_data[0] else [])]

    new_uids: list[int] = []
    emails_info: list[dict] = []
    combined_stats: dict = {}   # year-MM → {labo → {qty, total_ht}}

    for uid in candidate_uids:
        if uid in processed_uids:
            continue

        # Vérifier d'abord la structure (léger) avant de télécharger le message complet
        _, struct_data = mail.fetch(str(uid).encode(), "(BODYSTRUCTURE)")
        struct_str = str(struct_data[0] if struct_data else b"").lower()
        has_xlsx = "xlsx" in struct_str or "spreadsheetml" in struct_str or "excel" in struct_str
        if not has_xlsx:
            continue

        _, fetch_data = mail.fetch(str(uid).encode(), "(RFC822)")
        if not fetch_data or not fetch_data[0]:
            continue
        raw = fetch_data[0][1]
        msg = _email_lib.message_from_bytes(raw)

        # Chercher les pièces jointes XLSX dont le nom contient "justificatif"
        xlsx_bytes = None
        attach_name = ""
        for part in msg.walk():
            fn_raw = part.get_filename()
            if not fn_raw:
                continue
            # Décoder le nom encodé (iso-8859-1 / utf-8)
            decoded_parts = _decode_header(fn_raw)
            fn = "".join(
                (t.decode(enc or "utf-8") if isinstance(t, bytes) else t)
                for t, enc in decoded_parts
            )
            ct = part.get_content_type()
            is_xlsx = fn.lower().endswith(".xlsx") or "spreadsheetml" in ct or "excel" in ct
            if not is_xlsx:
                continue
            if "justificatif" not in fn.lower():
                continue
            payload = part.get_payload(decode=True)
            if payload and len(payload) > 1000:
                xlsx_bytes = payload
                attach_name = fn
                break

        if not xlsx_bytes:
            continue

        # Parser le fichier
        stats = _parse_grossiste_bytes(xlsx_bytes)
        if not stats:
            continue

        # Fusionner dans combined_stats (dict brut, avant conversion en listes)
        for mk, rows in stats.items():
            if mk not in combined_stats:
                combined_stats[mk] = {}
            for r in rows:
                acc = combined_stats[mk].setdefault(r["labo"], {"qty": 0, "total_ht": 0.0, "ca_brut": 0.0})
                acc["qty"]      += r["qty"]
                acc["total_ht"] += r["total_ht"]
                acc["ca_brut"]  += r.get("ca_brut", 0)

        new_uids.append(uid)
        subject_raw = msg.get("Subject", "")
        emails_info.append({
            "uid":     uid,
            "subject": subject_raw[:80],
            "date":    msg.get("Date", ""),
            "file":    attach_name,
            "months":  sorted(stats.keys()),
        })

    mail.logout()

    if not new_uids:
        return {
            "status":  "up_to_date",
            "emails_processed": 0,
            "processed_total":  len(processed_uids),
        }

    # Convertir combined_stats en format liste
    new_grossiste_stats = {
        mk: sorted(
            [{"labo": l, "qty": d["qty"], "total_ht": round(d["total_ht"], 2), "ca_brut": round(d.get("ca_brut", 0), 2)}
             for l, d in labos.items()],
            key=lambda r: r["labo"],
        )
        for mk, labos in sorted(combined_stats.items())
    }

    # Fusionner avec l'état existant et sauvegarder
    merged_stats = _merge_grossiste_stats(
        state.get("grossiste_month_stats") or {}, new_grossiste_stats
    )
    state["grossiste_month_stats"] = merged_stats
    state["grossiste_gmail_uids"] = sorted(processed_uids | set(new_uids))
    _patch_state_sync(user_id, state, user_token=user_token)

    total_q  = sum(r["qty"]      for rows in new_grossiste_stats.values() for r in rows)
    total_ht = sum(r["total_ht"] for rows in new_grossiste_stats.values() for r in rows)
    return {
        "status":                "done",
        "emails_processed":      len(new_uids),
        "emails":                emails_info,
        "months":                sorted(new_grossiste_stats.keys()),
        "labos":                 len({r["labo"] for rows in new_grossiste_stats.values() for r in rows}),
        "qty":                   total_q,
        "total_ht":              round(total_ht, 2),
        "grossiste_month_stats": merged_stats,
    }


# ── Parse PDF DIGI (upload direct) ───────────────────────────────────────────

@app.post("/parse/digi-pdf")
async def parse_digi_pdf(
    file: UploadFile = File(...),
    reset: bool = Form(False),
    authorization: str = Header(default=""),
):
    """Upload direct d'un PDF DIGI (invoice, RDP, presta) → extraction → digi_month_stats."""
    token = _extract_token(authorization)
    try:
        user_id = await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    name = (file.filename or "").lower()
    if not name.endswith(".pdf"):
        raise HTTPException(status_code=422, detail="Fichier PDF requis (.pdf)")

    pdf_bytes = await file.read()
    if len(pdf_bytes) < 200:
        raise HTTPException(status_code=422, detail="PDF trop court ou vide")

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        _executor,
        lambda: _parse_digi_pdf_sync(user_id, pdf_bytes, file.filename or "upload.pdf", user_token=token, reset=reset),
    )
    return result


def _parse_digi_pdf_sync(user_id: str, pdf_bytes: bytes, filename: str, user_token: str = "", reset: bool = False) -> dict:
    import os as _os, sys, tempfile
    from pathlib import Path as _Path

    sys.path.insert(0, _os.path.dirname(__file__))
    from pdf_extractor import extract_invoice_lines
    from run_job import (_compute_digi_month_stats, _merge_digi_stats,
                         _compute_escompte_stats, _merge_escompte_stats,
                         _compute_mdl_stats, _merge_mdl_stats)

    # Lecture/écriture de l'état via le TOKEN UTILISATEUR (rôle authenticated) —
    # comme le connecteur grossiste. La clé anon/service (_supa_key) n'a PAS le
    # droit d'UPDATE sur user_state (permission denied 42501 → 500), alors que
    # l'utilisateur peut écrire sa propre ligne (RLS auth.uid() = user_id).
    from supabase_client import _get_state_sync, _patch_state_sync

    # Écrire le PDF dans un fichier temp, extraire les lignes
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = _Path(tmp.name)
    try:
        provider = filename.rsplit(".", 1)[0][:80]
        lines    = extract_invoice_lines(tmp_path, provider, "")
    finally:
        tmp_path.unlink(missing_ok=True)

    if not lines:
        raise HTTPException(status_code=422, detail="Aucune donnée extractible depuis ce PDF — format non reconnu ou labo hors cible.")

    # Charger l'état courant (via token utilisateur)
    state = _get_state_sync(user_id, user_token=user_token) or {}

    escompte_lines = [l for l in lines if l.get("type") == "escompte"]
    mdl_lines      = [l for l in lines if l.get("type") == "mdl"]
    digi_lines     = [l for l in lines if l.get("type") not in ("escompte", "mdl")]

    # reset=True (ré-analyse) : on repart de bases vides → 1er fichier réinitialise.
    merged = {} if reset else (state.get("digi_month_stats") or {})
    if reset:
        state["digi_month_stats"] = merged
    if digi_lines:
        new_stats = _compute_digi_month_stats(digi_lines)
        merged    = _merge_digi_stats(merged, new_stats)
        state["digi_month_stats"] = merged

    merged_esc = {} if reset else (state.get("escompte_stats") or {})
    if reset:
        state["escompte_stats"] = merged_esc
    if escompte_lines:
        new_esc    = _compute_escompte_stats(escompte_lines)
        merged_esc = _merge_escompte_stats(merged_esc, new_esc)
        state["escompte_stats"] = merged_esc

    merged_mdl = {} if reset else (state.get("mdl_stats") or {})
    if reset:
        state["mdl_stats"] = merged_mdl
    if mdl_lines:
        new_mdl    = _compute_mdl_stats(mdl_lines)
        merged_mdl = _merge_mdl_stats(merged_mdl, new_mdl)
        state["mdl_stats"] = merged_mdl

    _patch_state_sync(user_id, state, user_token=user_token)

    months   = sorted(set(l.get("billing_date", "")[:7] for l in lines if l.get("billing_date")))
    n_rdp    = sum(1 for l in lines if l.get("type") == "rdp")
    n_presta = sum(1 for l in lines if l.get("type") == "presta")
    n_esc    = len(escompte_lines)
    n_mdl    = len(mdl_lines)
    n_prod   = len(lines) - n_rdp - n_presta - n_esc - n_mdl
    return {
        "status":           "done",
        "months":           months,
        "lines":            len(lines),
        "product_lines":    n_prod,
        "rdp_avoirs":       n_rdp,
        "presta_avoirs":    n_presta,
        "escompte_cerp":    n_esc,
        "mdl_cerp":         n_mdl,
        "digi_month_stats": merged,
        "escompte_stats":   merged_esc,
        "mdl_stats":        merged_mdl,
    }


# ── Parse XLSX FSE Banque (HTP+OI) ───────────────────────────────────────────
#
# Format export OSPHARM FSE Banque → XLSX avec colonnes :
@app.post("/import/rsf_history")
async def import_rsf_history(
    file:          UploadFile = File(...),
    labo:          str        = Form(...),
    year:          int        = Form(...),
    authorization: str        = Header(default=""),
):
    """Parse un PDF ou CSV CGV labo et upsert les taux RSF/RDP dans rsf_history."""
    token = _extract_token(authorization)
    try:
        await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    name = (file.filename or "").lower()
    if not any(name.endswith(e) for e in (".pdf", ".csv", ".xlsx")):
        raise HTTPException(status_code=422, detail="Fichier PDF, CSV ou XLSX requis")

    raw_bytes = await file.read()
    if len(raw_bytes) < 10:
        raise HTTPException(status_code=422, detail="Fichier trop court ou vide")

    fname = file.filename or "upload"
    loop = asyncio.get_event_loop()
    if name.endswith(".pdf"):
        result = await loop.run_in_executor(
            _executor,
            lambda: _import_rsf_history_sync(raw_bytes, labo, year, fname),
        )
    else:
        result = await loop.run_in_executor(
            _executor,
            lambda: _import_rsf_history_csv_sync(raw_bytes, labo, year, fname),
        )
    return result


@app.post("/import/rsf_exceptions")
async def import_rsf_exceptions(
    file:          UploadFile = File(...),
    labo:          str        = Form(...),
    year:          int        = Form(...),
    authorization: str        = Header(default=""),
):
    """Parse un fichier d'exceptions (xlsx/csv) et retourne les refs depuis rsf_history."""
    token = _extract_token(authorization)
    try:
        await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    name = (file.filename or "").lower()
    if not any(name.endswith(e) for e in (".pdf", ".csv", ".xlsx")):
        raise HTTPException(status_code=422, detail="Fichier PDF, CSV ou XLSX requis")

    raw_bytes = await file.read()
    if len(raw_bytes) < 10:
        raise HTTPException(status_code=422, detail="Fichier trop court ou vide")

    fname = file.filename or "upload"
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        _executor,
        lambda: _parse_rsf_exceptions_sync(raw_bytes, fname, labo, year),
    )
    return result


def _parse_rsf_exceptions_sync(raw_bytes: bytes, filename: str, labo: str, year: int) -> dict:
    """Extrait les CIP13 du fichier, les croise avec rsf_history et retourne les refs trouvées."""
    import urllib.parse
    from supabase_client import SUPA_URL, _supa_key

    # Extraire la liste des CIP13 depuis le fichier
    if filename.lower().endswith(".pdf"):
        import io, pdfplumber as _pdfplumber
        cips = []
        with _pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                cips += _re.findall(r"\b(\d{13})\b", text)
        cips = list(dict.fromkeys(cips))  # dédoublonnage ordre-préservé
    else:
        tabular = _parse_tabular_rows(raw_bytes, filename, labo=labo)
        cips = list(tabular.keys())

    if not cips:
        raise HTTPException(status_code=422, detail="Aucun CIP13 trouvé dans le fichier")

    tok = _supa_key()
    H = {"apikey": tok, "Authorization": f"Bearer {tok}"}
    labo_enc = urllib.parse.quote(labo)
    refs = []

    # Requête par chunks de 200 (limite URL safe)
    for i in range(0, len(cips), 200):
        chunk = cips[i:i+200]
        url = (f"{SUPA_URL}/rest/v1/rsf_history"
               f"?select=cip13,libelle,rsf_pct"
               f"&labo=eq.{labo_enc}&year=eq.{year}"
               f"&cip13=in.({','.join(chunk)})"
               f"&limit=200")
        req = urllib.request.Request(url, headers=H)
        with urllib.request.urlopen(req, timeout=30) as r:
            refs += json.loads(r.read())

    print(f"[import_rsf_exceptions] {labo} {year} → {len(refs)}/{len(cips)} CIP13 trouvés", flush=True)
    return {"refs": refs, "count": len(refs), "total_cips": len(cips)}


def _parse_tabular_rows(raw_bytes: bytes, filename: str, labo: str = "") -> "dict[str, tuple]":
    """Parse CSV ou XLSX → {cip13: (pfht, rsf, rdp, libelle)}.
    Colonnes attendues : CIP13 (obligatoire), LIBELLE (optionnel), RSF% (optionnel).
    La colonne RSF peut aussi s'appeler comme le labo (ex : "Teva").
    """
    import csv, io
    is_xlsx = filename.lower().endswith(".xlsx")
    _labo_lo = labo.strip().lower()

    def _pct(val) -> float:
        # openpyxl retourne 0.025 pour une cellule Excel formatée "2,50%" (stockée en décimal)
        if isinstance(val, (int, float)):
            v = float(val)
            if v != 0 and abs(v) < 1:
                v *= 100  # 0.025 → 2.5
            return -abs(v)
        v = str(val or "0").strip().rstrip("%").replace(",", ".").strip()
        return -abs(float(v)) if v else 0.0

    rows: dict[str, tuple] = {}

    if is_xlsx:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return rows
        # Première ligne = headers
        headers_raw = [str(c or "").strip() for c in all_rows[0]]
        headers_lo  = [h.lower() for h in headers_raw]
        cip_i = next((i for i, h in enumerate(headers_lo) if h in ("cip13", "cip")), None)
        rsf_i = next((i for i, h in enumerate(headers_lo) if "rsf" in h or "remise" in h or "taux" in h or "palier" in h or (_labo_lo and h == _labo_lo)), None)
        lib_i = next((i for i, h in enumerate(headers_lo) if "libelle" in h or (h.startswith("lib") and h != "libelle"[:len(h)] and "cip" not in h)), None)
        if lib_i is None:
            lib_i = next((i for i, h in enumerate(headers_lo) if "lib" in h and i != cip_i), None)
        print(f"[xlsx] headers={headers_raw} cip_i={cip_i} rsf_i={rsf_i} lib_i={lib_i}", flush=True)
        if rsf_i is not None and len(all_rows) > 1:
            sample = [repr(all_rows[i][rsf_i]) for i in range(1, min(4, len(all_rows)))]
            print(f"[xlsx] sample rsf col values: {sample}", flush=True)
        if cip_i is None:
            return rows
        for raw_row in all_rows[1:]:
            cip13 = str(raw_row[cip_i] or "").strip().replace(".0", "")
            if len(cip13) != 13 or not cip13.isdigit():
                continue
            rsf = 0.0
            if rsf_i is not None:
                try: rsf = _pct(raw_row[rsf_i])
                except (ValueError, TypeError): pass
            lib = str(raw_row[lib_i] or "").strip().upper() if lib_i is not None else cip13
            rows[cip13] = (0.0, rsf, 0.0, lib)
    else:
        # CSV
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                text = raw_bytes.decode(enc); break
            except UnicodeDecodeError:
                continue
        else:
            raise HTTPException(status_code=422, detail="Encodage CSV non reconnu")
        reader = csv.DictReader(io.StringIO(text))
        fieldnames = reader.fieldnames or []
        cip_col = next((f for f in fieldnames if f.strip().lower() in ("cip13", "cip")), None)
        rsf_col = next((f for f in fieldnames if any(k in f.strip().lower() for k in ("rsf", "remise", "taux", "palier")) or (_labo_lo and f.strip().lower() == _labo_lo)), None)
        lib_col = next((f for f in fieldnames if "libelle" in f.strip().lower() or ("lib" in f.strip().lower() and "cip" not in f.strip().lower())), None)
        if not cip_col:
            raise HTTPException(status_code=422, detail=f"Colonne CIP13 introuvable — colonnes : {fieldnames}")
        for row in reader:
            cip13 = (row.get(cip_col) or "").strip()
            if len(cip13) != 13 or not cip13.isdigit():
                continue
            rsf = 0.0
            if rsf_col:
                try: rsf = _pct(row.get(rsf_col, "0"))
                except ValueError: pass
            lib = (row.get(lib_col, "") or "").strip().upper() if lib_col else cip13
            rows[cip13] = (0.0, rsf, 0.0, lib)

    return rows


def _import_rsf_history_csv_sync(csv_bytes: bytes, labo: str, year: int, filename: str) -> dict:
    """Parse CSV ou XLSX CGV labo et upsert dans rsf_history."""
    import urllib.parse
    from supabase_client import SUPA_URL, _supa_key

    rows = _parse_tabular_rows(csv_bytes, filename, labo=labo)
    if not rows:
        raise HTTPException(status_code=422, detail="Aucun CIP13 valide trouvé dans le fichier")

    from normaliser import normaliser_libelle

    tok = _supa_key()
    HEADERS = {
        "apikey": tok, "Authorization": f"Bearer {tok}",
        "Content-Type": "application/json", "Prefer": "return=minimal",
    }

    labo_enc = urllib.parse.quote(labo)
    del_req = urllib.request.Request(
        f"{SUPA_URL}/rest/v1/rsf_history?labo=eq.{labo_enc}&year=eq.{year}",
        method="DELETE",
        headers={"apikey": tok, "Authorization": f"Bearer {tok}"},
    )
    with urllib.request.urlopen(del_req, timeout=30):
        pass

    rsf_data = [
        {"cip13": cip, "labo": labo, "year": year,
         "rsf_pct": rsf, "rdp_pct": rdp, "pfht": pfht,
         "libelle": normaliser_libelle(lib, cip)}
        for cip, (pfht, rsf, rdp, lib) in rows.items()
    ]
    chunk = 500
    for i in range(0, len(rsf_data), chunk):
        body = json.dumps(rsf_data[i:i+chunk]).encode()
        req  = urllib.request.Request(
            f"{SUPA_URL}/rest/v1/rsf_history",
            data=body, method="POST",
            headers={**HEADERS, "Prefer": "return=minimal"},
        )
        with urllib.request.urlopen(req, timeout=30):
            pass

    rpc_req = urllib.request.Request(
        f"{SUPA_URL}/rest/v1/rpc/refresh_references_from_rsf_history",
        data=json.dumps({"p_labo": labo}).encode(),
        method="POST",
        headers={"apikey": tok, "Authorization": f"Bearer {tok}", "Content-Type": "application/json"},
    )
    with urllib.request.urlopen(rpc_req, timeout=30) as resp:
        n_refs = json.loads(resp.read())
    print(f"  → references_pharmacie: {n_refs} lignes mises à jour depuis rsf_history", flush=True)

    try:
        from supabase_client import upload_file_sync
        upload_file_sync("admin", "rsf_history", f"{labo}_{year}_{filename.replace(' ', '_')}", csv_bytes, "text/csv")
    except Exception as e:
        print(f"[warn] MinIO upload failed: {e}", flush=True)

    from collections import Counter
    rsf_dist = dict(Counter(round(rsf, 4) for _, (_, rsf, _, _) in rows.items()).most_common(10))
    print(f"[import_rsf_history_csv] {labo} {year} → {len(rows)} CIP13 (libellés normalisés), rsf dist={rsf_dist}", flush=True)
    return {"count": len(rows), "labo": labo, "year": year, "rsf_dist": rsf_dist}


def _import_rsf_history_sync(pdf_bytes: bytes, labo: str, year: int, filename: str) -> dict:
    import io, pdfplumber as _pdfplumber
    import urllib.parse
    from supabase_client import SUPA_URL, _supa_key

    tok = _supa_key()

    # CIP13  libellé  PFHT€  REMISE%  RDP% (ou - quand pas de RDP)
    LINE_RE = _re.compile(
        r"(\d{13})\s+"
        r"(.+?)\s+"
        r"([\d]+[,.][\d]+)\s*€\s+"
        r"([\d]+(?:[,.][\d]+)?)\s*%\s+"
        r"(?:([\d]+(?:[,.][\d]+)?)\s*%|(-))"
    )
    # Viatris : CIP13  [libellé]  QTE_CARTON  QTE_FARDEAU  PFHT€  RSF_STD%  PRIX_STD€  [RSF_RIG%  PRIX_RIG€]
    # TAUX RIG (40 %) = RSF First, uniquement pour les produits du répertoire (palier 22 %).
    # Parser séquentiel (string ops + petits regex) pour éviter le backtracking catastrophique
    # du regex monolithique avec groupe optionnel lazy (?:(.+?)\s+)? qui prenait 53s sur 45 pages.
    # Extraction du texte : pymupdf (fitz) reconstruit les lignes par coordonnées Y — 53x plus
    # rapide que pdfplumber.extract_text() (1s vs 53s pour le PDF Viatris 45 pages, 2 Mo).
    _is_viatris = labo.strip().lower().startswith("viatris")
    _VP = r"[\d]+(?:[ \xa0]\d{3})?[,.]\d+"   # prix avec séparateur milliers optionnel

    def _viat_float(s: str) -> float:
        return float(s.replace("\xa0", "").replace(" ", "").replace(",", "."))

    def _parse_viatris_line(line: str) -> tuple | None:
        """Retourne (cip13, pfht, rsf, rsf_first, inline_lib) ou None."""
        if "€" not in line or "%" not in line:
            return None
        cip_m = _re.search(r"(?<!\d)(\d{13})(?!\d)", line)
        if not cip_m:
            return None
        cip13   = cip_m.group(1)
        cip_end = cip_m.end()

        # PFHT = juste avant le 1er € après le CIP13
        euro_pos = line.find("€", cip_end)
        if euro_pos == -1:
            return None
        pre_euro = line[cip_end:euro_pos]
        pfht_m = _re.search(r"(" + _VP + r")\s*$", pre_euro)
        if not pfht_m:
            return None
        pfht = _viat_float(pfht_m.group(1))

        # Filtre produits OTC/conseil : le dernier token AVANT le PFHT est un % décimal
        # (TVA type "10,0%" ou "5,5%") → format différent du répertoire (qui a QTE1 QTE2).
        # Ne pas utiliser "%" in inline_lib : les dosages comme "2% CREME" ou "0,2% 5ML"
        # contiennent aussi "%", mais leur DERNIER token avant PFHT est un entier (QTE_FARDEAU).
        pre_pfht = pre_euro[:pfht_m.start()].strip()
        last_tok = pre_pfht.rsplit(None, 1)[-1] if pre_pfht else ""
        if last_tok.endswith("%"):
            return None   # OTC/non-répertoire — TVA% immédiatement avant PFHT

        inline_lib = _re.sub(r"(\s+\d+)+\s*$", "", pre_pfht).strip()

        # RSF% et PRIX_STD€ après le PFHT€
        after_pfht = line[euro_pos + 1:]
        rsf_m = _re.search(r"([\d]+[,.]\d+)\s*%\s+" + _VP + r"\s*€", after_pfht)
        if not rsf_m:
            return None
        rsf = -abs(float(rsf_m.group(1).replace(",", ".")))

        # RIG% optionnel : même pattern après PRIX_STD€
        after_rsf = after_pfht[rsf_m.end():]
        rig_m = _re.search(r"([\d]+[,.]\d+)\s*%\s+" + _VP + r"\s*€", after_rsf)
        rsf_first = -abs(float(rig_m.group(1).replace(",", "."))) if rig_m else 0.0

        return (cip13, pfht, rsf, rsf_first, inline_lib)

    def _fitz_page_lines(page) -> list[str]:
        """Reconstruit les lignes du tableau Viatris en regroupant les mots par coordonnée Y."""
        from collections import defaultdict as _dd
        words = page.get_text("words")   # (x0,y0,x1,y1,word,block,line,word_no)
        by_y: dict = _dd(list)
        for x0, y0, _x1, _y1, word, *_ in words:
            by_y[round(y0 / 3) * 3].append((x0, word))
        lines = []
        for _y, row in sorted(by_y.items()):
            row.sort(key=lambda t: t[0])
            lines.append(" ".join(w for _, w in row))
        return lines

    # cip13 → (pfht, rsf, rdp, rsf_first, libelle)
    rows: dict[str, tuple] = {}

    if _is_viatris:
        import fitz as _fitz
        doc = _fitz.open(stream=io.BytesIO(pdf_bytes), filetype="pdf")
        for page in doc:
            prev_lib = ""
            for line in _fitz_page_lines(page):
                result = _parse_viatris_line(line)
                if result:
                    cip13, pfht, rsf, rsf_first, inline_lib = result
                    # Si libellé inline court (<20 car) → c'est la colonne INFORMATION
                    # ("BUD", "Switch", "Hors répertoire"), le vrai nom est dans prev_lib
                    if prev_lib and len(inline_lib) < 20:
                        lib_raw = (prev_lib + (" " + inline_lib if inline_lib else "")).strip()
                    else:
                        lib_raw = inline_lib or prev_lib
                    rows[cip13] = (pfht, rsf, 0.0, rsf_first, lib_raw.upper())
                    prev_lib = ""
                else:
                    stripped = line.strip()
                    if stripped and "€" not in line and not _re.search(r"\d{13}", line):
                        prev_lib = stripped
                    else:
                        prev_lib = ""
        doc.close()
    else:
        with _pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    m = LINE_RE.search(line)
                    if m:
                        cip13   = m.group(1)
                        libelle = m.group(2).strip().upper()
                        pfht    = float(m.group(3).replace(",", "."))
                        rsf     = -abs(float(m.group(4).replace(",", ".")))
                        rdp     = -abs(float(m.group(5).replace(",", "."))) if m.group(5) else 0.0
                        rows[cip13] = (pfht, rsf, rdp, 0.0, libelle)

    if not rows:
        raise HTTPException(status_code=422, detail="Aucun CIP13 trouvé dans le PDF")

    from normaliser import normaliser_libelle

    HEADERS = {
        "apikey":        tok,
        "Authorization": f"Bearer {tok}",
        "Content-Type":  "application/json",
        "Prefer":        "resolution=merge-duplicates,return=minimal",
    }

    # 1. Supprimer les anciennes lignes labo+year (remplacement propre)
    labo_enc = urllib.parse.quote(labo)
    del_req = urllib.request.Request(
        f"{SUPA_URL}/rest/v1/rsf_history?labo=eq.{labo_enc}&year=eq.{year}",
        method="DELETE",
        headers={"apikey": tok, "Authorization": f"Bearer {tok}"},
    )
    with urllib.request.urlopen(del_req, timeout=30):
        pass

    # 2. Insérer toutes les nouvelles lignes avec libellé normalisé
    rsf_data = [
        {"cip13": cip, "labo": labo, "year": year,
         "rsf_pct": rsf, "rdp_pct": rdp, "rsf_first_pct": rsf_first, "pfht": pfht,
         "libelle": normaliser_libelle(lib, cip)}
        for cip, (pfht, rsf, rdp, rsf_first, lib) in rows.items()
    ]
    chunk = 500
    for i in range(0, len(rsf_data), chunk):
        body = json.dumps(rsf_data[i:i+chunk]).encode()
        req  = urllib.request.Request(
            f"{SUPA_URL}/rest/v1/rsf_history",
            data=body, method="POST",
            headers={**HEADERS, "Prefer": "return=minimal"},
        )
        with urllib.request.urlopen(req, timeout=30):
            pass

    # 3. Rafraîchir references_pharmacie depuis rsf_history (millésime le plus récent par CIP13)
    rpc_req = urllib.request.Request(
        f"{SUPA_URL}/rest/v1/rpc/refresh_references_from_rsf_history",
        data=json.dumps({"p_labo": labo}).encode(),
        method="POST",
        headers={"apikey": tok, "Authorization": f"Bearer {tok}", "Content-Type": "application/json"},
    )
    with urllib.request.urlopen(rpc_req, timeout=30) as resp:
        n_refs = json.loads(resp.read())
    print(f"  → references_pharmacie: {n_refs} lignes mises à jour depuis rsf_history", flush=True)

    # 4. Upload PDF dans MinIO pour archive
    try:
        from supabase_client import upload_file_sync
        _safe_filename = filename.replace(" ", "_")
        upload_file_sync("admin", "rsf_history", f"{labo}_{year}_{_safe_filename}", pdf_bytes, "application/pdf")
    except Exception as e:
        print(f"[warn] MinIO upload failed: {e}", flush=True)

    print(f"[import_rsf_history] {labo} {year} → {len(rows)} CIP13 (libellés normalisés)", flush=True)
    return {"count": len(rows), "labo": labo, "year": year}


#   Date (DD/MM/YYYY) | Libellé | Montant (TTC, euros)
#
# Libellé format labo : "VIR BIOGARAN - 9006671913 - 2000065455 - Emetteur 300030154000020476361"
# Montant en TTC (virements bancaires incluent TVA si applicable).
# RDP (R2) : hors TVA → montant = HT = TTC
# Presta (R3) : TVA 20% → montant TTC

class ParseFseBankBody(BaseModel):
    storage_path: str  # chemin dans bucket 'fse-bank', ex: "user_id/ts_filename.xlsx"

@app.post("/parse/fse-bank")
async def parse_fse_bank(
    body: ParseFseBankBody,
    authorization: str = Header(default=""),
):
    """Parse le XLSX d'export FSE Banque (HTP+OI) et sauvegarde fse_month_stats."""
    token = _extract_token(authorization)
    try:
        user_id = await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        _executor,
        lambda: _parse_fse_bank_sync(user_id, body.storage_path),
    )
    return result


@app.post("/parse/fse-bank-xlsx")
async def parse_fse_bank_xlsx(
    file: UploadFile = File(...),
    reset: bool = Form(False),
    authorization: str = Header(default=""),
):
    """Upload direct multipart du XLSX FSE Banque (sans MinIO, auth S3 inaccessible côté
    navigateur) → fse_month_stats. Même approche que /parse/grossiste-xlsx et /parse/digi-pdf."""
    token = _extract_token(authorization)
    try:
        user_id = await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    name = (file.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls")):
        raise HTTPException(status_code=422, detail="Fichier XLSX requis (.xlsx/.xls)")
    xlsx_bytes = await file.read()
    if len(xlsx_bytes) < 200:
        raise HTTPException(status_code=422, detail="Fichier trop court ou vide")

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        _executor, lambda: _parse_fse_bank_sync(user_id, xlsx_bytes=xlsx_bytes, user_token=token, reset=reset)
    )


def _parse_fse_bank_sync(user_id: str, storage_path: str = "", xlsx_bytes: bytes = None, user_token: str = "", reset: bool = False) -> dict:
    import io, re as _re, datetime as _dt
    import openpyxl

    from supabase_client import SUPA_URL, _supa_key
    _tok = _supa_key()
    HEADERS = {"apikey": _tok, "Authorization": f"Bearer {_tok}"}

    # ── Labos génériqueurs connus ──────────────────────────────────────────────
    _LABO_KEYS = [
        ("BIOGARAN", "BIOGARAN"), ("TEVA", "TEVA"), ("VIATRIS", "VIATRIS"),
        ("MYLAN", "VIATRIS"), ("SANDOZ", "SANDOZ"), ("ZENTIVA", "ZENTIVA"),
        ("ARROW", "ARROW"), ("CRISTERS", "CRISTERS"), ("ZYDUS", "ZYDUS"),
        ("EG LABO", "EG LABO"), ("EG LABS", "EG LABO"), ("EVOLUPHARM", "EVOLUPHARM"),
        ("RANBAXY", "RANBAXY"), ("AUROBINDO", "AUROBINDO"), ("INTAS", "INTAS"),
        ("ALMUS", "ALMUS"), ("QUALIMED", "QUALIMED"),
        # Dépositaires qui virent au nom du labo
        ("MOVIANTO", "MOVIANTO"), ("ALLOGA", "ALLOGA"), ("CEGEDIM", "CEGEDIM"),
        # Grossistes qui virent leurs ristournes directement
        ("CERP", "CERP"), ("COOPERATION PHARMACEUTIQUE", "CERP"), ("CPF", "CERP"),
    ]

    def _identify_labo(libelle: str) -> str | None:
        """Extrait le labo depuis le libellé VIR {NOM} - {REF} - ..."""
        lib = libelle.upper().strip()
        m = _re.match(r'VIR\s+(.+?)\s+-\s', lib)
        if not m:
            # Chercher directement un nom de labo dans l'intégralité du libellé
            for key, canon in _LABO_KEYS:
                if key in lib:
                    return canon
            return None
        name_part = m.group(1)
        for key, canon in _LABO_KEYS:
            if name_part.startswith(key) or key in name_part:
                return canon
        return None

    def _extract_all_refs(libelle: str) -> list[str]:
        """Numéros de facture potentiels : groupes 8-14 chiffres + tokens
        alphanumériques (ex. '24BO24121001288')."""
        refs = [s for s in _re.findall(r'\b(\d{8,14})\b', libelle) if len(s) <= 14]
        for tok in _re.findall(r'\b([A-Z0-9]{8,20})\b', libelle.upper()):
            if any(c.isalpha() for c in tok) and any(c.isdigit() for c in tok):
                refs.append(tok)
        return refs

    def _classify_transfer(libelle: str, ref: str) -> str:
        """Classifie le virement en 'r2' (RDP) ou 'r3' (prestation coopérative)."""
        text = (libelle + ' ' + ref).upper()
        r2_kw = ['RDP', ' R2 ', '-R2-', 'R2-', '-R2', 'REMISE FIN', 'AVOIR', 'REDUCTI', 'RED FIN']
        r3_kw = ['PRESTA', 'COOP', ' R3 ', '-R3-', 'R3-', '-R3', 'PREST', 'COOPERAT', 'PRESTAT']
        if any(kw in text for kw in r2_kw):
            return 'r2'
        if any(kw in text for kw in r3_kw):
            return 'r3'
        return 'r3'

    def _parse_date(val) -> str | None:
        """Convertit DD/MM/YYYY ou datetime en YYYY-MM-DD."""
        if isinstance(val, (_dt.date, _dt.datetime)):
            return val.strftime("%Y-%m-%d")
        s = str(val or "").strip()
        m = _re.match(r'(\d{2})/(\d{2})/(\d{4})', s)
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else None

    def _parse_amount(val) -> float | None:
        """Parse '2 161,74' ou '2161.74' ou '2 994€' → float."""
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val or "").replace('\xa0', '').replace(' ', '').replace('€', '').replace(',', '.').strip()
        try:
            return float(s) if s else None
        except ValueError:
            return None

    # 1. Obtenir les octets du XLSX : soit fournis en multipart (upload direct, sans MinIO),
    #    soit téléchargés depuis Storage (ancien chemin via bucket 'fse-bank').
    if xlsx_bytes is None:
        dl_url = f"{SUPA_URL}/storage/v1/object/fse-bank/{storage_path}"
        req = urllib.request.Request(dl_url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=30) as r:
            xlsx_bytes = r.read()

    # 2. Parser le XLSX
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Charger toutes les lignes d'un coup (max 5000 lignes)
    all_rows = list(ws.iter_rows(max_row=5000, values_only=True))

    # ── Détection des colonnes ─────────────────────────────────────────────────
    # Approche 1 : chercher la ligne d'en-tête par nom de colonne
    _DATE_KW   = ('date',)
    _LIB_KW    = ('libell', 'libellé', 'description', 'opération', 'operation', 'désignation')
    _AMT_KW    = ('montant', 'crédit', 'credit', 'débit', 'debit', 'valeur', 'amount')
    hdr_idx    = -1
    col_date, col_lib, col_amt = 0, 1, 2  # defaults

    for ri, row in enumerate(all_rows[:15]):
        cells = [str(c or '').lower().strip() for c in row]
        has_date = any(any(k in c for k in _DATE_KW) for c in cells)
        has_lib  = any(any(k in c for k in _LIB_KW)  for c in cells)
        if has_date and has_lib:
            hdr_idx = ri
            for ci, c in enumerate(cells):
                if any(k in c for k in _DATE_KW) and 'mise' not in c and 'update' not in c:
                    col_date = ci
                if any(k in c for k in _LIB_KW):
                    col_lib  = ci
                if any(k in c for k in _AMT_KW):
                    col_amt  = ci
            break

    # Approche 2 : détection par contenu — chercher la colonne avec "VIR " dans les cellules
    if hdr_idx < 0:
        for ri, row in enumerate(all_rows[:30]):
            for ci, c in enumerate(row):
                s = str(c or '').upper().strip()
                if s.startswith('VIR ') and len(s) > 6:
                    # data starts here; header probably 1 row above
                    hdr_idx  = max(0, ri - 1)
                    col_lib  = ci
                    col_date = max(0, ci - 1)   # date probablement à gauche
                    col_amt  = ci + 1            # montant probablement à droite
                    break
            if hdr_idx >= 0:
                break

    data_start = hdr_idx + 2  # 0-indexed; openpyxl min_row est 1-indexed donc +1

    # ── Table ref→labo depuis digi_month_stats ──────────────────────────────────
    # Les numéros de factures RDP/presta extraits des PDFs Digi sont les mêmes
    # que ceux présents dans les libellés de virements FSE. On construit un
    # dictionnaire de lookup pour identifier le labo quand le nom est absent.
    req_state = urllib.request.Request(
        f"{SUPA_URL}/rest/v1/user_state?user_id=eq.{user_id}&select=state_json&limit=1",
        headers=HEADERS,
    )
    try:
        with urllib.request.urlopen(req_state, timeout=15) as _r:
            _rows = json.loads(_r.read())
        _state_pre = (_rows[0]["state_json"] if _rows else {}) or {}
    except Exception:
        _state_pre = {}

    _digi_stats = _state_pre.get("digi_month_stats") or {}
    ref_to_labo: dict[str, str] = {}   # ref_number → canonical_labo
    ref_to_type: dict[str, str] = {}   # ref_number → 'r2' | 'r3'
    norm_ref_index: list[tuple[str, str, str]] = []   # (norm_ref, labo, type)
    # amount_to_candidates : montant_cents → [(labo, month_key, type)]
    # Permet d'identifier le labo quand le libellé ne contient ni nom ni ref
    amount_to_candidates: dict[int, list[tuple[str, str, str]]] = {}
    for _mk_d, _arr in _digi_stats.items():
        for _row in (_arr or []):
            _labo_d = _row.get("labo", "")
            for _ref in (_row.get("facture_refs") or []):
                _ref_s = str(_ref).strip()
                if _ref_s and len(_ref_s) >= 6:
                    _typ_r = 'r2' if (_row.get("rdp_total") or 0) > 0 else 'r3'
                    ref_to_labo[_ref_s] = _labo_d
                    ref_to_type[_ref_s] = _typ_r
                    _nr = _re.sub(r'[^A-Z0-9]', '', _ref_s.upper())
                    if len(_nr) >= 7:
                        norm_ref_index.append((_nr, _labo_d, _typ_r))
            # Presta TTC → r3 ; RDP → r2
            for _field, _typ in [("presta_total_ttc", "r3"), ("rdp_total", "r2")]:
                _amt = _row.get(_field) or 0
                if _amt > 0:
                    _cents = round(_amt * 100)
                    amount_to_candidates.setdefault(_cents, []).append((_labo_d, _mk_d, _typ))
    norm_ref_index.sort(key=lambda x: -len(x[0]))   # n° les plus longs d'abord
    print(f"  → Lookup facture refs : {len(ref_to_labo)} entrées ({len(norm_ref_index)} normalisés) · {len(amount_to_candidates)} montants distincts depuis digi_month_stats")

    # ── Parcours des lignes de données ─────────────────────────────────────────
    acc: dict[str, dict] = {}
    virements_list: list[dict] = []
    unmatched_virements: list[dict] = []
    row_num         = 0
    n_ref_matched   = 0
    n_amt_matched   = 0
    skipped_lib: list[str] = []

    for row in all_rows[data_start:]:
        if not any(c for c in row):
            continue
        n = len(row)
        date_str = _parse_date(row[col_date] if col_date < n else None)
        libelle  = str(row[col_lib] if col_lib < n else '') or ''
        amount   = _parse_amount(row[col_amt] if col_amt < n else None)

        if (amount is None or amount <= 0) and col_amt + 1 < n:
            amount = _parse_amount(row[col_amt + 1])

        if not date_str or not libelle:
            continue
        if amount is None or amount <= 0:
            continue
        if not libelle.upper().strip().startswith('VIR '):
            continue

        labo  = _identify_labo(libelle)
        refs  = _extract_all_refs(libelle)
        vtype = _classify_transfer(libelle, refs[0] if refs else "")

        # Fallback 1 : cross-référence par numéro de facture dans les libellés
        if not labo and refs:
            for _r in refs:
                if _r in ref_to_labo:
                    labo  = ref_to_labo[_r]
                    vtype = ref_to_type.get(_r, vtype)
                    n_ref_matched += 1
                    break

        # Fallback 1b : n° de facture Digi en sous-chaîne du libellé normalisé
        # (n° souvent noyé dans le libellé, ex. ".../INV/24BO24121001288...").
        if not labo and norm_ref_index:
            _norm_lib = _re.sub(r'[^A-Z0-9]', '', libelle.upper())
            for _nref, _nlabo, _ntyp in norm_ref_index:
                if _nref in _norm_lib:
                    labo, vtype = _nlabo, _ntyp
                    n_ref_matched += 1
                    break

        # Fallback 2 : cross-référence par montant TTC (presta/rdp Digi)
        # Cherche d'abord le mois correspondant, sinon prend le premier candidat
        if not labo:
            _cents = round(amount * 100)
            _cands = amount_to_candidates.get(_cents)
            if _cands:
                _match = next((c for c in _cands if c[1] == mk), _cands[0])
                labo   = _match[0]
                vtype  = _match[2]
                n_amt_matched += 1

        if not labo:
            if len(skipped_lib) < 20:
                skipped_lib.append(libelle[:80])
            unmatched_virements.append({
                "date": date_str, "mois": mk, "montant_ttc": amount,
                "libelle": libelle[:120], "refs": refs[:3],
            })
            continue

        # Normaliser le nom de labo avec la table canonique
        for key, canon in _LABO_KEYS:
            if key in labo.upper():
                labo = canon
                break

        mk = date_str[:7]
        acc.setdefault(mk, {}).setdefault(labo, {"montant_ttc": 0.0, "r2_ttc": 0.0, "r3_ttc": 0.0, "count": 0, "refs": []})
        acc[mk][labo]["montant_ttc"] += amount
        acc[mk][labo]["r2_ttc"]      += amount if vtype == 'r2' else 0.0
        acc[mk][labo]["r3_ttc"]      += amount if vtype == 'r3' else 0.0
        acc[mk][labo]["count"]       += 1
        for _r in refs[:3]:
            if _r not in acc[mk][labo]["refs"]:
                acc[mk][labo]["refs"].append(_r)
        virements_list.append({
            "date": date_str, "mois": mk, "labo": labo, "type": vtype,
            "montant_ttc": amount, "refs": refs[:3], "libelle": libelle[:100],
        })
        row_num += 1

    print(f"  → {row_num} parsés · {n_ref_matched} via ref · {n_amt_matched} via montant · {len(unmatched_virements)} non identifiés")
    if skipped_lib:
        print(f"  → Libellés non reconnus : {skipped_lib[:5]}")

    if not acc:
        # Collecter infos de debug : premières lignes + colonnes détectées
        sample = [[str(c)[:25] for c in r if c] for r in all_rows[max(0,hdr_idx):hdr_idx+5] if any(c for c in r)]
        debug  = (
            f"Colonnes détectées : date={col_date}, libellé={col_lib}, montant={col_amt}. "
            f"Ligne header={hdr_idx}. "
            f"Premières lignes : {sample[:3]}. "
            f"Libellés VIR non reconnus : {skipped_lib[:5]}"
        )
        raise HTTPException(status_code=422, detail=debug)

    fse_stats = {
        mk: {
            labo: {
                "montant_ttc": round(d["montant_ttc"], 2),
                "r2_ttc":      round(d["r2_ttc"],      2),
                "r3_ttc":      round(d["r3_ttc"],      2),
                "count":       d["count"],
                "refs":        d["refs"][:10],
            }
            for labo, d in labos.items()
        }
        for mk, labos in sorted(acc.items())
    }

    # 3. Fusionner avec l'état existant
    req2 = urllib.request.Request(
        f"{SUPA_URL}/rest/v1/user_state?user_id=eq.{user_id}&select=state_json&limit=1",
        headers=HEADERS,
    )
    with urllib.request.urlopen(req2, timeout=15) as r:
        rows2 = json.loads(r.read())
    state = (rows2[0]["state_json"] if rows2 else {}) or {}
    existing = {} if reset else (state.get("fse_month_stats") or {})

    # Migration ancien format (dict-of-lists) → dict-of-dicts
    merged: dict = {}
    for mk, lab_data in existing.items():
        if isinstance(lab_data, list):
            merged[mk] = {r["labo"]: dict(r) for r in lab_data}
        else:
            merged[mk] = dict(lab_data)

    for mk, new_labos in fse_stats.items():
        if mk not in merged:
            merged[mk] = new_labos
        else:
            for labo, nd in new_labos.items():
                if labo in merged[mk]:
                    ex = merged[mk][labo]
                    ex["montant_ttc"] = round(ex.get("montant_ttc", 0) + nd["montant_ttc"], 2)
                    ex["r2_ttc"]      = round(ex.get("r2_ttc",      0) + nd["r2_ttc"],      2)
                    ex["r3_ttc"]      = round(ex.get("r3_ttc",      0) + nd["r3_ttc"],      2)
                    ex["count"]      += nd["count"]
                    ex["refs"]        = (ex.get("refs", []) + nd["refs"])[:20]
                else:
                    merged[mk][labo] = dict(nd)
    state["fse_month_stats"] = merged

    # Écriture avec le TOKEN UTILISATEUR (RLS : MAJ de sa propre ligne ; la clé anon
    # est en lecture seule sur user_state → 401 en écriture).
    from supabase_client import _patch_state_sync
    _patch_state_sync(user_id, state, user_token=user_token)

    months = sorted(fse_stats)
    total_ttc = sum(d["montant_ttc"] for labos in fse_stats.values() for d in labos.values())
    return {
        "status":          "done",
        "months":          months,
        "rows_parsed":     row_num,
        "total_ttc":       round(total_ttc, 2),
        "fse_month_stats":     merged,
        "virements":           virements_list[:500],
        "unmatched_virements": unmatched_virements[:100],
    }


# ── Conn test (async wrapper) ──────────────────────────────────────────────────

async def _run_conn_test_async(user_id: str, connector: str, creds: dict, user_token: str = ""):
    loop = asyncio.get_event_loop()
    try:
        if connector == "digipharmacie":
            # 1. Tenter curl_cffi directement sur Render (rapide, pas de runner)
            try:
                await loop.run_in_executor(_executor, lambda: _test_digi_curl_only(creds))
                await save_user_creds(user_id, connector, creds["user"], creds["pass"], True, user_token)
                await patch_conn_test(user_id, connector, True, "Connexion réussie")
                return
            except RuntimeError as e:
                # Mauvais credentials → fail immédiat, pas besoin du runner
                await patch_conn_test(user_id, connector, False, str(e))
                return
            except Exception:
                pass  # Cloudflare bloque Render → fallback runner self-hosted

            # 2. Cloudflare bloque → dispatch vers runner self-hosted (IP résidentielle)
            await save_user_creds(user_id, connector, creds["user"], creds["pass"], False, user_token)
            await _dispatch_gh_conn_test(user_id, connector)
            return

        if connector in ("ospharm", "concentrateur"):
            await save_user_creds(user_id, connector, creds["user"], creds["pass"], True, user_token)
            await patch_conn_test(user_id, connector, True,
                                  "Identifiants enregistrés — vérifiés au prochain lancement",
                                  user_token=user_token)
            return

        if connector == "gmail":
            try:
                import imaplib
                mail = imaplib.IMAP4_SSL("imap.gmail.com", timeout=20)
                mail.login(creds["user"], creds["pass"])
                mail.logout()
            except Exception as e:
                await patch_conn_test(user_id, connector, False,
                                      f"Connexion IMAP échouée : {e}", user_token=user_token)
                return
            await save_user_creds(user_id, connector, creds["user"], creds["pass"], True, user_token)
            await patch_conn_test(user_id, connector, True, "Connexion IMAP réussie", user_token=user_token)
            return

        try:
            await loop.run_in_executor(_executor, lambda: _test_connector(connector, creds, user_id))
            await save_user_creds(user_id, connector, creds["user"], creds["pass"], True)
            await patch_conn_test(user_id, connector, True, "Connexion réussie")
        except Exception as e:
            await patch_conn_test(user_id, connector, False, str(e))

    except Exception as _top_err:
        print(f"  [conn-test] erreur inattendue {connector}: {_top_err}", flush=True)
        try:
            await patch_conn_test(user_id, connector, False, f"Erreur serveur : {_top_err}")
        except Exception:
            pass


async def _dispatch_gh_conn_test(user_id: str, connector: str):
    """Déclenche test_connector.yml sur GitHub Actions (self-hosted, IP non bloquée)."""
    if not GH_TOKEN:
        await patch_conn_test(user_id, connector, False,
                              "GH_TOKEN manquant sur le serveur — contacter l'admin")
        return

    url  = f"https://api.github.com/repos/{GH_REPO}/actions/workflows/{GH_TEST_WORKFLOW}/dispatches"
    body = json.dumps({
        "ref": "master",
        "inputs": {"user_id": user_id, "connector": connector},
    }).encode()
    req  = urllib.request.Request(url, data=body, method="POST", headers={
        "Authorization": f"Bearer {GH_TOKEN}",
        "Accept":        "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "Content-Type":  "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            print(f"  [gh-test] HTTP {r.status} — test {connector} pour {user_id[:8]}")
    except Exception as e:
        print(f"  [gh-test] ERREUR: {e}")
        await patch_conn_test(user_id, connector, False,
                              f"Impossible de lancer le workflow GitHub: {e}")


# ── Test connector (synchronous, called from executor) ─────────────────────────

def _test_digi_curl_only(creds: dict):
    """Test Digipharmacie via curl_cffi uniquement — RuntimeError si mauvais credentials,
    autre Exception si Cloudflare bloque (le caller dispatch alors vers le runner self-hosted)."""
    from test_connector import test_digi_curl
    test_digi_curl(creds)


def _test_connector(connector: str, creds: dict, user_id: str = ""):
    if connector in ("ospharm", "concentrateur"):
        asyncio.set_event_loop(asyncio.new_event_loop())
        if connector == "ospharm":
            from test_connector import test_ospharm
            test_ospharm(creds)
        else:
            from test_connector import test_concentrateur
            test_concentrateur(creds)
    elif connector == "digipharmacie":
        # Chemin rapide : curl_cffi en-processus (~5-10s, pas de navigateur)
        try:
            from test_connector import test_digi_curl
            test_digi_curl(creds)
            return  # succès
        except RuntimeError:
            raise  # mauvais credentials
        except Exception as curl_err:
            if GH_TOKEN or os.environ.get("PROXY_URL"):
                # Proxy configuré mais toujours bloqué — camoufox subprocess ne servira à rien
                raise RuntimeError(f"Cloudflare bloque malgré le proxy : {curl_err}")
            pass  # pas de proxy → fallback subprocess camoufox

        # Fallback : subprocess camoufox avec hard timeout 180s
        _run_digi_test_subprocess(user_id, creds)


def _run_digi_test_subprocess(user_id: str, creds: dict):
    import subprocess
    import sys
    env = dict(os.environ)
    env["CONNECTOR"]  = "digipharmacie"
    env["USER_ID"]    = user_id
    env["DIGI_USER"]  = creds.get("user", "")
    env["DIGI_PASS"]  = creds.get("pass", "")
    try:
        proc = subprocess.run(
            [sys.executable, "test_connector.py"],
            env=env,
            timeout=180,
            capture_output=True,
            text=True,
        )
    except subprocess.TimeoutExpired:
        raise RuntimeError(
            "Timeout (>180s) — Digipharmacie inaccessible depuis ce serveur "
            "(Cloudflare bloque les IPs Render). Contactez le support."
        )
    if proc.returncode != 0:
        out = (proc.stdout + "\n" + proc.stderr).strip()
        lines = [l.strip() for l in out.splitlines() if l.strip()]
        raise RuntimeError(lines[-1] if lines else "Test Digipharmacie échoué")


# ── Background job ─────────────────────────────────────────────────────────────

async def _run_job_async(job_id: str, user_id: str, connector: str, job_key: str, creds: dict):
    loop = asyncio.get_event_loop()

    def progress(msg: str):
        if job_id in _jobs:
            _jobs[job_id]["message"] = msg

    try:
        result = await loop.run_in_executor(
            _executor,
            lambda: _scrape(connector, user_id, creds, progress),
        )
        if isinstance(result, (tuple, list)):
            rows       = result[0]
            file_url   = result[1] if len(result) > 1 else ""
            period_start = result[2] if len(result) > 2 else ""
            period_end   = result[3] if len(result) > 3 else ""
        else:
            rows, file_url, period_start, period_end = result, "", "", ""
        # Pour OSPHARM : compacter à {cip13, qty, libelle} avant stockage Supabase
        # (réduit ~5 Mo → ~400 Ko ; ospharmRowsToCsvData() sur le front gère les deux formats)
        stored_rows = _compact_osp_rows(rows) if connector == "ospharm" else rows
        msg = f"{len(rows)} lignes extraites"
        _jobs[job_id].update({
            "status":   "done",
            "message":  msg,
            "rows":     stored_rows,
            "total":    len(rows),
            "file_url": file_url,
        })
        await patch_job_status(user_id, job_key, "done", msg, stored_rows, file_url,
                               period_start=period_start, period_end=period_end)
    except Exception as e:
        _jobs[job_id].update({"status": "error", "message": str(e), "error": str(e)})
        await patch_job_status(user_id, job_key, "error", str(e), [])


def _compact_osp_rows(rows: list[dict]) -> list[dict]:
    """Convertit les lignes OSPHARM brutes (24 cols) en {cip13, qty, libelle}.
    Réduit ~5 Mo → ~400 Ko pour le stockage dans Supabase.
    Même logique que ospharmRowsToCsvData() côté frontend.
    """
    if not rows:
        return []

    def _n(k):
        s = (k or "").lower()
        for a, b in [("é","e"),("è","e"),("ê","e"),("à","a"),("ù","u"),("î","i"),("ô","o")]:
            s = s.replace(a, b)
        return _re.sub(r"[^a-z0-9]", "", s)

    keys = list(rows[0].keys())
    cip_k = next((k for k in keys if _n(k) == "codeean"), None) or \
            next((k for k in keys if any(p in _n(k) for p in ("cip", "ean", "acl"))), None)
    qty_k = next((k for k in keys if _n(k) == "quantite"), None) or \
            next((k for k in keys if any(p in _n(k) for p in ("qte", "qty"))
                  and "n1" not in _n(k) and "evo" not in _n(k)), None)
    lib_k = next((k for k in keys if _n(k) == "libelleproduit"), None) or \
            next((k for k in keys if "produit" in _n(k)), None) or \
            next((k for k in keys if "libelle" in _n(k)), None)

    if not cip_k or not qty_k:
        return rows  # fallback: renvoyer les données brutes si colonnes non trouvées

    result = []
    for r in rows:
        raw = _re.sub(r"\D", "", str(r.get(cip_k) or ""))
        cip13 = raw if len(raw) == 13 else ("340000" + raw if len(raw) == 7 else None)
        try:
            qty = float(str(r.get(qty_k) or 0).replace(",", "."))
        except (ValueError, TypeError):
            qty = 0.0
        if not cip13 or qty <= 0:
            continue
        result.append({
            "cip13":   cip13,
            "qty":     qty,
            "libelle": str(r.get(lib_k) or "").strip() if lib_k else "",
        })
    return result


def _scrape(connector: str, user_id: str, creds: dict, progress):
    if connector == "digipharmacie":
        # async camoufox — asyncio.run() crée sa propre boucle
        from scraper import run_scraper
        return run_scraper(creds, progress)
    elif connector == "ospharm":
        asyncio.set_event_loop(asyncio.new_event_loop())
        from run_job_ospharm import run_ospharm
        return run_ospharm(creds, progress, user_id=user_id)
    raise RuntimeError(f"Connecteur inconnu : {connector}")


# ── Exploration Digipharmacie Espaces clients (endpoint temporaire) ────────────

@app.get("/explore/digi-espace-client")
async def explore_digi_espace_client(
    background_tasks: BackgroundTasks,
    authorization: str = Header(default=""),
):
    """Lance l'exploration en arrière-plan. Résultat dans state_json.digi_espace_client_explore."""
    token = _extract_token(authorization)
    try:
        user_id = await verify_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    background_tasks.add_task(
        lambda: _executor.submit(_explore_digi_espace_client_sync, user_id)
    )
    return {"status": "started", "check": "state_json.digi_espace_client_explore"}


def _explore_digi_espace_client_sync(user_id: str) -> dict:
    import asyncio as _asyncio
    from supabase_client import SUPA_URL, _supa_key
    _tok = _supa_key()

    creds = {}
    try:
        url = f"{SUPA_URL}/rest/v1/user_state?user_id=eq.{user_id}&select=connectors&limit=1"
        req = urllib.request.Request(url, headers={"apikey": _tok, "Authorization": f"Bearer {_tok}"})
        with urllib.request.urlopen(req, timeout=15) as r:
            rows = json.loads(r.read())
        conns = (rows[0].get("connectors") or {}) if rows else {}
        cred  = conns.get("digipharmacie", {})
        creds = {"user": cred.get("user", ""), "pass": cred.get("pass", "")}
    except Exception as e:
        return {"error": f"Impossible de lire les credentials: {e}"}

    if not creds.get("user"):
        return {"error": "Pas de credentials Digipharmacie en base"}

    _asyncio.set_event_loop(_asyncio.new_event_loop())
    return _asyncio.get_event_loop().run_until_complete(_explore_async(creds, user_id))


async def _explore_async(creds: dict, user_id: str) -> dict:
    import os as _os
    from camoufox.async_api import AsyncCamoufox
    from supabase_client import SUPA_URL, _supa_key
    _tok = _supa_key()

    BASE = "https://app.digipharmacie.fr"
    PROXY_URL = _os.environ.get("PROXY_URL", "")
    api_calls = []
    pages_visited = []

    proxy_cfg = None
    if PROXY_URL:
        import urllib.parse as _up
        _p = _up.urlparse(PROXY_URL)
        proxy_cfg = {"server": f"{_p.scheme}://{_p.hostname}:{_p.port}",
                     "username": _p.username or "", "password": _p.password or ""}

    # curl_cffi login
    session_cookies: dict = {}
    try:
        from curl_cffi import requests as cffi_requests
        proxy_kw = {"proxy": PROXY_URL} if PROXY_URL else {}
        session = cffi_requests.Session(impersonate="chrome124")
        r = session.get(f"{BASE}/login/", headers={"Accept": "text/html,*/*", "Accept-Language": "fr-FR,fr;q=0.9"},
                        timeout=25, allow_redirects=True, **proxy_kw)
        print(f"[explore] curl GET /login/ → {r.status_code} body={len(r.text)}b")
        csrf = session.cookies.get("csrftoken", "")
        print(f"[explore] csrf={'ok' if csrf else 'MANQUANT'}")
        if csrf:
            for ep in ["/api/v1/auth/login/", "/api/auth/login/"]:
                rp = session.post(f"{BASE}{ep}",
                                  json={"email": creds["user"], "password": creds["pass"]},
                                  headers={"Accept": "application/json", "Content-Type": "application/json",
                                           "X-CSRFToken": csrf, "Referer": f"{BASE}/login/", "Origin": BASE},
                                  timeout=15, allow_redirects=False, **proxy_kw)
                if rp.status_code == 200:
                    session_cookies = dict(session.cookies)
                    break
                if rp.status_code in (400, 401):
                    return {"error": "Identifiants DIGIPHARMACIE incorrects"}
            if not session_cookies:
                rp = session.post(f"{BASE}/login/",
                                  data={"email": creds["user"], "password": creds["pass"], "csrfmiddlewaretoken": csrf},
                                  headers={"Content-Type": "application/x-www-form-urlencoded", "Referer": f"{BASE}/login/"},
                                  timeout=15, allow_redirects=True, **proxy_kw)
                if "/login" not in rp.url:
                    session_cookies = dict(session.cookies)
    except Exception:
        pass

    from scraper import _ensure_camoufox
    await _ensure_camoufox()

    async with AsyncCamoufox(headless=True, geoip=False) as browser:
        ctx  = await browser.new_context(**({"proxy": proxy_cfg} if proxy_cfg else {}))
        if session_cookies:
            await ctx.add_cookies([{"name": k, "value": v, "domain": "app.digipharmacie.fr",
                                    "path": "/", "sameSite": "Lax"} for k, v in session_cookies.items()])

        page = await ctx.new_page()
        page.on("pageerror", lambda e: None)
        page.set_default_timeout(30_000)

        async def on_response(resp):
            if "/api/" in resp.url and resp.status < 400:
                try:
                    body = await resp.json()
                    api_calls.append({"url": resp.url.replace(BASE, ""), "status": resp.status,
                                      "body": body if not isinstance(body, list) else body[:3],
                                      "count": len(body) if isinstance(body, list) else None})
                except Exception:
                    pass
        page.on("response", on_response)

        # Vérifier la session
        await page.goto(f"{BASE}/", timeout=30_000)
        await page.wait_for_load_state("networkidle", timeout=10_000)
        print(f"[explore] camoufox / → {page.url}  title={await page.title()}")
        pages_visited.append({"label": "home", "url": page.url})

        # Login camoufox si pas de cookies
        if "/login" in page.url and not session_cookies:
            cf_kw = ("just a moment", "checking", "verifying", "cloudflare")
            for _ in range(20):
                title = (await page.title()).lower()
                if not any(k in title for k in cf_kw):
                    break
                await page.wait_for_timeout(3_000)
            await page.fill("input[type=email], input[name=email]", creds["user"])
            await page.fill("input[type=password]", creds["pass"])
            await page.keyboard.press("Enter")
            await page.wait_for_load_state("networkidle", timeout=20_000)

        pages_visited.append({"label": "after_login", "url": page.url})

        # Naviguer vers Espaces clients
        await page.goto(f"{BASE}/achat/espaces-clients/", timeout=20_000)
        await page.wait_for_load_state("networkidle", timeout=10_000)
        pages_visited.append({"label": "espaces_clients", "url": page.url})

        html     = await page.content()
        all_links = await page.evaluate("""() =>
            Array.from(document.querySelectorAll('a')).map(a => ({
                text: a.textContent.trim().slice(0, 80), href: a.href
            })).filter(a => a.text && a.href.includes('digipharmacie'))
        """)
        await ctx.close()

    result = {"pages_visited": pages_visited, "api_calls": api_calls,
              "links": all_links[:30], "html": html[:8000]}

    # Sauvegarder en base
    try:
        url  = f"{SUPA_URL}/rest/v1/user_state?user_id=eq.{user_id}&select=state_json&limit=1"
        req  = urllib.request.Request(url, headers={"apikey": _tok, "Authorization": f"Bearer {_tok}"})
        with urllib.request.urlopen(req, timeout=15) as r:
            rows = json.loads(r.read())
        state = rows[0]["state_json"] if rows else {}
        state["digi_espace_client_explore"] = result
        body = json.dumps({"state_json": state}).encode()
        req2 = urllib.request.Request(f"{SUPA_URL}/rest/v1/user_state?user_id=eq.{user_id}",
                                      data=body, method="PATCH",
                                      headers={"apikey": _tok, "Authorization": f"Bearer {_tok}",
                                               "Content-Type": "application/json", "Prefer": "return=minimal"})
        with urllib.request.urlopen(req2, timeout=15): pass
    except Exception as e:
        result["save_error"] = str(e)

    return result
